# -*- coding: utf-8 -*-
"""
관세 인보이스(CI)와 원산지 증명서(CO) PDF 데이터 추출 및 통합 프로그램
Invoice No 기준으로 두 서류 데이터를 결합하여 엑셀로 저장합니다.

[추출 원칙]
- PDF 한 박스(셀) = 엑셀 한 셀. 박스 안의 숫자·문자는 모두 한 셀에 넣고, 줄 단위로 나누지 않음.
- 박스가 여러 줄이면 셀 값 내부에 \\n으로 유지하여 엑셀에서 한 셀에 여러 줄로 표시.
- 각 박스 안의 숫자·문자는 모두 추출하여 누락 없이 엑셀에 기록.

필요 라이브러리 설치 (아래 명령어를 터미널에서 실행):
  # 기본 (Tesseract OCR 사용, Python 3.14 등 모든 버전 지원)
  pip install pdfplumber pdf2image pandas openpyxl pillow opencv-python pytesseract
  # + Windows에서 Tesseract 엔진: https://github.com/UB-Mannheim/tesseract/wiki 에서 설치 후 PATH 등록

  # 선택: PaddleOCR 사용 시 (Python 3.9~3.13)
  pip install paddlepaddle paddleocr
"""

import os
import re
from pathlib import Path

import cv2
import numpy as np
import pandas as pd
import pdf2image
import pdfplumber

# PaddleOCR 선택적 사용 (미설치 시 Tesseract로 대체)
try:
    from paddleocr import PaddleOCR
    _HAS_PADDLE = True
except ImportError:
    PaddleOCR = None
    _HAS_PADDLE = False

try:
    import pytesseract
    _HAS_TESSERACT = True
    # Windows: Tesseract가 PATH에 없으면 기본 설치 경로 사용
    if os.name == "nt":
        _tesseract_exe = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
        if os.path.isfile(_tesseract_exe):
            pytesseract.pytesseract.tesseract_cmd = _tesseract_exe
except ImportError:
    pytesseract = None
    _HAS_TESSERACT = False


def _safe_print(msg):
    """한글/특수문자 포함 메시지를 콘솔에 안전하게 출력 (Windows cp949 등)."""
    import sys
    try:
        print(msg)
    except UnicodeEncodeError:
        enc = getattr(sys.stdout, "encoding", None) or "utf-8"
        print(msg.encode(enc, errors="replace").decode(enc))


# =============================================================================
# 전역 설정
# =============================================================================

# PaddleOCR 엔진 (한 번만 초기화하여 재사용)
_ocr_engine = None


def get_ocr_engine():
    """
    OCR 엔진 반환: PaddleOCR이 있으면 사용, 없으면 None (Tesseract 경로 사용).
    Tesseract는 extract_text_from_co_image에서 engine이 None일 때 호출됨.
    """
    global _ocr_engine
    if not _HAS_PADDLE:
        return None
    if _ocr_engine is None:
        _ocr_engine = PaddleOCR(use_angle_cls=True, lang="en", show_log=False)
    return _ocr_engine


# =============================================================================
# 유형 1: Commercial Invoice 추출 (pdfplumber)
# =============================================================================


def extract_invoice_header(page):
    """
    인보이스 페이지 상단에서 'Invoice No'와 'Invoice Date'를 추출합니다.
    텍스트 라인을 순회하며 키워드 매칭으로 값을 찾습니다.
    """
    invoice_no = None
    invoice_date = None
    text = page.extract_text()
    if not text:
        return invoice_no, invoice_date

    lines = text.split("\n")
    for line in lines:
        line_lower = line.lower()
        # "Invoice No" 또는 "Invoice No." 패턴
        if "invoice no" in line_lower and invoice_no is None:
            # 콜론 또는 공백 뒤의 값 추출
            match = re.search(r"invoice\s*no\.?\s*:?\s*([A-Za-z0-9\-]+)", line, re.I)
            if match:
                invoice_no = match.group(1).strip()
        if "invoice date" in line_lower and invoice_date is None:
            match = re.search(
                r"invoice\s*date\.?\s*:?\s*(\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4}|\d{4}[/\-\.]\d{1,2}[/\-\.]\d{1,2})",
                line,
                re.I,
            )
            if match:
                invoice_date = match.group(1).strip()

    return invoice_no, invoice_date


def merge_continuation_rows(table_rows, key_col_indices):
    """
    PDF 한 박스가 여러 행으로 나뉘어 추출된 경우(한 셀 = 여러 줄)를 한 행으로 합칩니다.
    - key_col_indices: 키 컬럼 인덱스 (Order No, Product Code). 이 컬럼이 있으면 '주 행'.
    - 연속 행: 키 컬럼은 비어 있고 다른 컬럼에만 값이 있는 행.
    - 방향: (1) 다음 행이 연속이면 현재 행에 이어붙임. (2) 현재 행이 주 행이고 이전 행이 연속이면 이전 행을 현재 행 앞에 \\n으로 붙임.
    - 원칙: PDF 한 박스 = 엑셀 한 셀. 박스 안 여러 줄은 한 셀 내에서 \\n으로 유지.
    """
    if not table_rows or not key_col_indices:
        return table_rows

    def is_key_empty(row, indices):
        return all(
            i < len(row) and (row[i] is None or not str(row[i]).strip())
            for i in indices
        )

    def has_other_content(row, indices):
        return any(
            k < len(row) and row[k] is not None and str(row[k]).strip()
            for k in range(len(row)) if k not in indices
        )

    merged = []
    i = 0
    while i < len(table_rows):
        row = list(table_rows[i])
        # 1) 다음 행들이 연속(키 비어 있음)이면 현재 행에 이어붙임
        j = i + 1
        while j < len(table_rows):
            next_row = list(table_rows[j])
            if not is_key_empty(next_row, key_col_indices):
                break
            if not has_other_content(next_row, key_col_indices):
                break
            for k in range(max(len(row), len(next_row))):
                if k >= len(row):
                    row.append(None)
                nv = next_row[k] if k < len(next_row) else None
                nv_str = str(nv).strip() if nv else ""
                if not nv_str:
                    continue
                cur = row[k]
                cur_str = str(cur).strip() if cur else ""
                row[k] = (cur_str + "\n" + nv_str) if cur_str else nv_str
            j += 1
        merged.append(row)
        i = j

    # 2) 연속 행(키 비어 있음)을 다음 주 행(키 있음) 앞에 \n으로 붙임 (한 박스 = 한 셀)
    # 단, 서브헤더 행(No., Description, Category, of Origin 등만 있는 행)은 버퍼에 넣지 않음
    SUBHEADER_PATTERN = re.compile(
        r"^(no\.?|description|category|quality|of\s*origin|price|gender|brand|product)$",
        re.I,
    )
    # 버퍼 셀 값이 이 키워드면 데이터 셀 앞에 붙이지 않음 (헤더 유출 방지)
    HEADER_CELL_KEYWORDS = frozenset(
        ("no.", "no", "description", "category", "quality", "of origin", "price", "gender", "brand", "product")
    )

    def looks_like_subheader_row(row, key_indices):
        """셀 값이 전부 No., Description, Category 등 서브헤더 키워드면 True."""
        for k in range(len(row)):
            if k in key_indices:
                continue
            val = row[k]
            if not val or not str(val).strip():
                continue
            s = str(val).strip().lower()
            if not SUBHEADER_PATTERN.match(s) and len(s) > 3:
                return False  # 일반 데이터가 하나라도 있으면 서브헤더 아님
        return True

    def is_header_cell_value(s):
        """셀 값이 헤더 키워드(No., Description, Category 등)만 있으면 True."""
        if not s or not str(s).strip():
            return True
        t = str(s).strip().lower()
        return t in HEADER_CELL_KEYWORDS or SUBHEADER_PATTERN.match(t)

    result = []
    buffer = []  # 연속 행들 (키 비어 있음)
    for row in merged:
        row = list(row)
        if is_key_empty(row, key_col_indices) and has_other_content(row, key_col_indices):
            if looks_like_subheader_row(row, key_col_indices):
                buffer = []  # 서브헤더 행은 버퍼 비우고 스킵 (다음 주 행에 붙이지 않음)
                continue
            buffer.append(row)
            continue
        # 주 행: 버퍼 내용을 이 행 앞에 붙임 (한 박스 = 한 셀). 헤더 키워드(No., Description 등)는 붙이지 않음.
        if buffer:
            for k in range(max(len(row), *[len(b) for b in buffer])):
                if k >= len(row):
                    row.extend([None] * (k - len(row) + 1))
                cur = row[k]
                cur_str = str(cur).strip() if cur else ""
                for bi, br in enumerate(buffer):
                    if k >= len(br):
                        continue
                    bv = br[k]
                    bv_str = str(bv).strip() if bv else ""
                    if not bv_str:
                        continue
                    bv_lower = bv_str.lower()
                    # 헤더 셀(No., Description 등)은 기존 데이터 앞에만 붙이지 않음. 빈 셀에는 유지(Product Category 등)
                    if bv_lower in HEADER_CELL_KEYWORDS or SUBHEADER_PATTERN.match(bv_str):
                        if cur_str:
                            continue  # 이미 데이터가 있으면 헤더를 앞에 붙이지 않음 → Crocs만 유지
                        row[k] = bv_str  # 빈 셀에는 버퍼 값 유지 (Category 등 실제 데이터)
                        cur_str = bv_str
                        continue
                    # 중복 방지: 버퍼 값이 이미 주 행 값에 포함되면 붙이지 않음
                    if cur_str and bv_str in cur_str:
                        continue
                    # 중복 방지: 같은 문구가 반복되면 한 번만 (KINGMAKER...CO.,LTDKINGMAKER... → 한 번만)
                    if cur_str and bv_str and len(bv_str) >= 10:
                        if bv_str in cur_str or cur_str.endswith(bv_str[: min(15, len(bv_str))]):
                            continue
                    if cur_str and cur_str in bv_str:
                        row[k] = bv_str
                        cur_str = bv_str
                        continue
                    # 앞이 조각·뒤가 색상코드(예: Char/Blk)면 공백으로 이어 한 문장으로
                    if cur_str and len(cur_str) < 30 and re.match(r"^[A-Za-z]+\s*/\s*[A-Za-z]+$", cur_str.strip()):
                        row[k] = (bv_str + " " + cur_str).strip()
                    else:
                        row[k] = (bv_str + "\n" + cur_str) if cur_str else bv_str
                    cur_str = str(row[k]).strip() if row[k] else ""
            buffer = []
        result.append(row)
    # 남은 버퍼(마지막이 연속 행)는 이전 result 행에 이어붙임 (중복·헤더 유출 방지 동일 적용)
    if buffer and result:
        last = result[-1]
        for br in buffer:
            for k in range(max(len(last), len(br))):
                if k >= len(last):
                    last.append(None)
                bv = br[k] if k < len(br) else None
                bv_str = str(bv).strip() if bv else ""
                if not bv_str:
                    continue
                bv_lower = bv_str.lower()
                if bv_lower in HEADER_CELL_KEYWORDS or SUBHEADER_PATTERN.match(bv_str):
                    continue
                cur = last[k]
                cur_str = str(cur).strip() if cur else ""
                if cur_str and bv_str in cur_str:
                    continue
                if cur_str and len(bv_str) >= 10 and (bv_str in cur_str or cur_str.endswith(bv_str[: min(15, len(bv_str))])):
                    continue
                last[k] = (cur_str + "\n" + bv_str) if cur_str else bv_str
    elif buffer:
        result.extend(buffer)
    return result


def merge_description_rows(table_rows, desc_col_index):
    """
    테이블에서 'Description' 등 한 항목이 여러 행으로 나뉜 경우,
    연속된 빈 셀이 있는 행을 이전 행의 Description과 합쳐서 하나의 행으로 만듭니다.
    desc_col_index: Description 컬럼의 인덱스 (0-based).
    (merge_continuation_rows로 전체 컬럼 병합 시 이 함수는 보조로만 사용 가능)
    """
    if not table_rows or desc_col_index < 0:
        return table_rows

    merged = []
    i = 0
    while i < len(table_rows):
        row = list(table_rows[i])
        if len(row) <= desc_col_index:
            merged.append(row)
            i += 1
            continue

        desc_cell = str(row[desc_col_index]).strip() if row[desc_col_index] else ""
        j = i + 1
        while j < len(table_rows):
            next_row = list(table_rows[j])
            if len(next_row) <= desc_col_index:
                break
            next_desc = (
                str(next_row[desc_col_index]).strip() if next_row[desc_col_index] else ""
            )
            front_empty = all(
                not str(c).strip() for c in next_row[:desc_col_index]
            )
            if front_empty and next_desc:
                desc_cell = desc_cell + "\n" + next_desc
                j += 1
            else:
                break

        row[desc_col_index] = desc_cell
        merged.append(row)
        i = j

    return merged


def normalize_invoice_table_columns(row, expected_len=9):
    """
    테이블 한 행을 표준 컬럼 수에 맞춤.
    기대 컬럼: Invoice No, Invoice Date, Order No, Product Code, Description, HTS No, Qty, Unit Price, Total
    """
    row = list(row)
    while len(row) < expected_len:
        row.append("")
    return row[:expected_len]


# 표준 인보이스 컬럼명과 매칭할 헤더 키워드 (소문자)
# 원칙: PDF 한 박스 = 엑셀 한 셀. 각 박스 안 숫자·문자는 모두 한 셀에 추출.
_INVOICE_HEADER_KEYWORDS = {
    "Invoice No": ["invoice no", "invoice no.", "inv no", "inv. no", "invoice number"],
    "Invoice Date": ["invoice date", "inv date", "date of invoice", "invoice date"],
    "Order No": ["order no", "order no.", "order number", "order #", "po no", "p.o. no", "order", "order ref"],
    "Product Code": ["product code", "item code", "article", "article no", "part no", "sku", "code", "item no", "style", "model"],
    "Description": ["description", "item description", "product description", "goods description", "품목", "desc", "item", "product name"],
    "Gender": ["gender"],
    "Brand Description": ["brand description", "brand"],
    "Product Category": ["product category", "category"],
    "Quality": ["quality"],
    "HTS No": ["hts", "hs code", "hs no", "tariff", "hts no", "harmonized", "customs", "hts code"],
    "Country of Origin": ["country of origin", "origin", "country"],
    "Qty": ["qty", "quantity", "quantities", "수량", "qty.", "order qty", "units"],
    "UOM": ["uom", "unit of measure", "pair", "pcs", "ea"],
    "Unit Price": ["unit price", "price", "unit price (", "unit price)", "단가", "price/unit", "unit value"],
    "Total": ["total", "amount", "extended", "합계", "total amount", "ext. amount", "line total", "value", "amount total"],
    "Manufacturer": ["manufacturer"],
}


def _fallback_col_mapping(data_rows):
    """데이터 행들에서 컬럼 수가 가장 많은 행 기준으로 순서 기반 컬럼 매핑 (모든 행의 컬럼 커버)."""
    if not data_rows:
        return {}
    max_cols = max(len(list(r)) for r in data_rows)
    fallback_names = [
        "Order No", "Product Code", "Description", "Gender", "Brand Description", "Product Category",
        "Quality", "HTS No", "Country of Origin", "Qty", "UOM", "Unit Price", "Total", "Manufacturer",
    ]
    return {i: fallback_names[i] for i in range(min(max_cols, len(fallback_names)))}


def _forward_fill_merged_cells(records, fill_keys=None):
    """
    PDF 병합 셀 때문에 위쪽 행에만 값이 있고 아래 행은 빈 경우, 이전 행 값으로 채움.
    fill_keys: 채울 컬럼 (기본: Order No, Product Code, Product Category, Unit Price).
    현재 행이 품목 데이터(Description/Qty/Total 중 하나라도 있음)일 때만 이전 값으로 채움.
    """
    if fill_keys is None:
        fill_keys = ("Order No", "Product Code", "Product Category", "Unit Price")
    if len(records) < 2:
        return
    for j in range(1, len(records)):
        rec = records[j]
        prev = records[j - 1]
        has_item_data = any(
            rec.get(k) and str(rec.get(k)).strip()
            for k in ("Description", "Qty", "Unit Price", "Total", "HTS No")
        )
        if not has_item_data:
            continue
        for key in fill_keys:
            if key not in rec:
                continue
            cur_val = rec.get(key)
            is_empty = cur_val is None or (isinstance(cur_val, str) and not cur_val.strip())
            prev_val = prev.get(key)
            has_prev = prev_val is not None and (not isinstance(prev_val, str) or prev_val.strip())
            if is_empty and has_prev:
                rec[key] = prev_val


def _back_fill_merged_cells(records, fill_keys=None):
    """
    서브헤더 행(예: HTS No에 "No."만 있는 행)은 Order No가 비어 있음.
    그 다음 행에 Order No가 있으면, 서브헤더 행에 다음 행 값을 채움 (back-fill).
    """
    if fill_keys is None:
        fill_keys = ("Order No", "Product Code", "Product Category")
    if len(records) < 2:
        return
    for j in range(len(records) - 1, -1, -1):
        rec = records[j]
        # 현재 행이 서브헤더처럼 보이는지 (Order No 비어 있고, HTS No가 "No." 등)
        order_empty = not (rec.get("Order No") and str(rec.get("Order No")).strip())
        hts_val = str(rec.get("HTS No") or "").strip().lower()
        looks_like_subheader = order_empty and (hts_val == "no." or not rec.get("Description") and not rec.get("Qty"))
        if not looks_like_subheader:
            continue
        if j + 1 >= len(records):
            continue
        next_rec = records[j + 1]
        for key in fill_keys:
            next_val = next_rec.get(key)
            if next_val is not None and (not isinstance(next_val, str) or next_val.strip()):
                rec[key] = next_val


def _fix_product_code_description_split(rec):
    """
    PDF 셀 경계로 인해 Description 앞글자가 Product Code 셀에 들어간 경우 보정.
    예: Product Code "Cl\\n211936-3TC", Description "assic..." → PC "211936-3TC", Desc "Classic..."
    예: Product Code "Ech\\n211981-082", Description "o Duck..." → PC "211981-082", Desc "Echo Duck..."
    """
    pc = rec.get("Product Code")
    desc = rec.get("Description")
    if not pc or not isinstance(pc, str):
        return
    pc = pc.strip()
    desc = (desc or "").strip() if isinstance(desc, str) else ""

    # 1) Product Code에 줄바꿈이 있고, 맨 앞 연속의 짧은 알파벳 줄(2~5자)만 Description 앞으로 이동
    # 예: "Cl\\n211936-3TC" → PC "211936-3TC", Desc "Classic...". "Cl\\nTh\\n206990-100..." → "Cl"만 이동
    if "\n" in pc:
        lines = [ln.strip() for ln in pc.split("\n") if ln.strip()]
        prefix_for_desc = []
        code_lines = []
        seen_digit_line = False
        for ln in lines:
            is_short_alpha = bool(re.match(r"^[A-Za-z]{2,5}$", ln) and not re.search(r"\d", ln))
            if not seen_digit_line and is_short_alpha and len(prefix_for_desc) < 2:
                prefix_for_desc.append(ln)
            else:
                if re.search(r"\d", ln):
                    seen_digit_line = True
                code_lines.append(ln)
        if prefix_for_desc and code_lines:
            rec["Product Code"] = "\n".join(code_lines)
            prepend = " ".join(prefix_for_desc)
            rec["Description"] = (prepend + " " + desc).strip() if desc else prepend
            return

    # 2) Product Code 끝이 짧은 알파벳(예: Croc, N)이고 Description이 이어지는 단어(예: band, BA)로 시작하는 경우
    if not desc:
        return
    # "212756-001 Croc" + "band Gum Clog" → PC "212756-001", Desc "Crocband Gum Clog"
    m = re.match(r"^(.+?)\s+([A-Za-z]{1,4})$", pc)
    if m:
        rest, suffix = m.group(1), m.group(2)
        if re.search(r"\d", rest) and (
            desc.lower().startswith("band")
            or desc.lower().startswith("ba ")
            or (len(suffix) <= 2 and desc.lower().startswith("a "))
        ):
            rec["Product Code"] = rest.strip()
            rec["Description"] = (suffix + desc).strip()
            return
    # "212727-90H N" + "BA Echo..." → PC "212727-90H", Desc "NBA Echo..." 또는 "N BA Echo..."
    if pc.endswith(" N") and desc.upper().startswith("BA "):
        rec["Product Code"] = pc[:-2].strip()
        rec["Description"] = ("N" + desc).strip()
        return

    # 3) Description만 셀에 들어와 앞글자가 잘린 경우 보정 (Product Code에서 가져오지 못했을 때)
    DESC_FRAGMENT_PREFIX = (
        ("assic", "Cl"),   # Classic
        ("ermoplastic", "Th"),  # Thermoplastic
        ("ocband", "Cr"),  # Crocband
    )
    for fragment, prefix in DESC_FRAGMENT_PREFIX:
        if desc.lower().startswith(fragment):
            rec["Description"] = (prefix + desc).strip()
            return
    if desc.startswith("o ") or desc.startswith("o Duck"):
        rec["Description"] = ("Ech" + desc).strip()
        return


def _merge_same_order_no_rows(records):
    """
    동일 Order No로 두 행이 나뉜 경우 병합 (첫 행에는 Order No·Product Code만, 다음 행에 나머지).
    예: R21 [Order No, Product Code, 빈칸...] + R22 [Order No, Product Code, Description, Qty, ...] → 한 행으로.
    """
    if not records or len(records) < 2:
        return
    i = 0
    while i < len(records):
        rec = records[i]
        order_no = str(rec.get("Order No") or "").strip()
        has_desc_or_qty = bool(
            (rec.get("Description") and str(rec.get("Description")).strip())
            or (rec.get("Qty") is not None and rec.get("Qty") != "" and str(rec.get("Qty")).strip())
            or (rec.get("Total") is not None and rec.get("Total") != "" and str(rec.get("Total")).strip())
        )
        if not order_no or has_desc_or_qty:
            i += 1
            continue
        if i + 1 >= len(records):
            i += 1
            continue
        next_rec = records[i + 1]
        next_order = str(next_rec.get("Order No") or "").strip()
        if next_order != order_no:
            i += 1
            continue
        next_has = bool(
            (next_rec.get("Description") and str(next_rec.get("Description")).strip())
            or (next_rec.get("Qty") is not None and next_rec.get("Qty") != "" and str(next_rec.get("Qty")).strip())
            or (next_rec.get("Total") is not None and next_rec.get("Total") != "" and str(next_rec.get("Total")).strip())
        )
        if not next_has:
            i += 1
            continue
        for key in rec:
            if key in ("Order No", "Invoice No", "Invoice Date"):
                continue
            cur_val = rec.get(key)
            is_empty = cur_val is None or (isinstance(cur_val, str) and not str(cur_val).strip())
            if is_empty:
                next_val = next_rec.get(key)
                if next_val is not None and (not isinstance(next_val, str) or str(next_val).strip()):
                    rec[key] = next_val
        records.pop(i + 1)
        # 다음에도 같은 Order No가 올 수 있으므로 i 증가하지 않고 다시 검사


def _deduplicate_repeated_phrase(text):
    """
    박스 내 동일 문구가 중복으로 붙은 경우 한 번만 남김.
    예: KINGMAKER III(VN)FOOTWEARCO.,LTDKINGMAKER III(VN)FOOTWEAR → KINGMAKER III(VN)FOOTWEARCO.,LTD
    """
    if not text or not isinstance(text, str) or len(text) < 20:
        return text
    s = text.strip()
    # 앞쪽에서 시작하는 문구가 뒤쪽에 다시 나오면 두 번째 등장 제거 (긴 문구부터 시도)
    for length in range(min(60, len(s) // 2), 7, -1):
        if length > len(s) - 5:
            continue
        phrase = s[: length]
        if phrase in ("", " ", ".", ",") or phrase.count(" ") == len(phrase):
            continue
        idx = s.find(phrase, length)
        if idx == -1:
            continue
        # s = phrase + 중간 + phrase 형태면 phrase + 중간 만 남김
        return (s[: idx] + s[idx + length :]).strip()
    return s


def _header_cell_to_standard_name(cell_text):
    """헤더 셀 텍스트가 어떤 표준 컬럼에 해당하는지 반환. 없으면 None."""
    if not cell_text:
        return None
    s = str(cell_text).strip().lower()
    for std_name, keywords in _INVOICE_HEADER_KEYWORDS.items():
        for kw in keywords:
            if kw in s:
                return std_name
    return None


def _find_header_row_index(table):
    """
    ​테이블(2D 리스트)에서 헤더 행 인덱스를 찾습니다.
    표준 컬럼명과 매칭되는 셀이 가장 많은 행을 헤더로 선택합니다.
    """
    if not table or len(table) < 2:
        return 0
    best_idx = 0
    best_score = 0
    for idx, row in enumerate(table[: min(10, len(table))]):  # 상위 10행만 검사
        score = sum(1 for c in row if _header_cell_to_standard_name(c) is not None)
        if score > best_score:
            best_score = score
            best_idx = idx
    return best_idx


def _build_col_index_to_standard_name(header_row):
    """
    헤더 행(리스트)을 보고 각 인덱스 → 표준 컬럼명 매핑을 반환.
    {0: "Order No", 1: "Description", ...} 형태.
    """
    mapping = {}
    for i, cell in enumerate(header_row):
        name = _header_cell_to_standard_name(cell)
        if name:
            mapping[i] = name
    return mapping


def _extract_table_with_strategy(page, table_settings):
    """지정한 table_settings로 첫 번째 테이블 추출. 실패 시 None."""
    try:
        tables = page.find_tables(table_settings=table_settings)
        if tables:
            return tables[0].extract()
    except Exception:
        pass
    return None


def _score_table_as_item_table(table):
    """
    테이블이 품목(라인 아이템) 테이블인지 점수로 평가.
    - 헤더 행에 Description, Qty, Unit Price, Total 등이 많을수록 높은 점수.
    - 데이터 행 수가 적당할수록(5~200) 가산, 너무 많으면(전체 페이지 테이블) 감점.
    """
    if not table or len(table) < 2:
        return -1
    best_header_score = 0
    for idx, row in enumerate(table[: min(5, len(table))]):
        mapping = _build_col_index_to_standard_name(row)
        item_keys = {"Description", "Qty", "Unit Price", "Total", "Product Code", "Order No", "HTS No"}
        score = sum(1 for std in mapping.values() if std in item_keys)
        if score > best_header_score:
            best_header_score = score
    data_row_count = max(0, len(table) - 1)
    # 행 수 보정: 5~150행이면 가산, 200행 넘으면 감점 (페이지 전체를 테이블로 잡은 경우)
    if data_row_count > 200:
        row_penalty = (data_row_count - 200) // 10
    elif 5 <= data_row_count <= 150:
        row_penalty = -2  # 가산
    else:
        row_penalty = 0
    return best_header_score * 10 + row_penalty - min(data_row_count // 50, 5)


def _extract_invoice_table_from_page(page):
    """
    한 페이지에서 인보이스 **품목** 테이블을 추출합니다.
    여러 테이블 후보 중 헤더에 Description/Qty/Unit Price 등이 있는 테이블을 선택하고,
    헤더 행 자동 감지 후 컬럼 매핑하여 데이터 행만 반환합니다.
    반환: (header_mapping: dict, data_rows: list) 또는 (None, []) 실패 시.
    """
    candidates = []  # (table, col_mapping, data_rows)

    # 1) extract_tables()로 모든 테이블 후보 수집
    try:
        all_tables = page.extract_tables()
        if all_tables:
            for table in all_tables:
                if not table or len(table) < 2:
                    continue
                header_idx = _find_header_row_index(table)
                header_row = table[header_idx]
                data_rows = table[header_idx + 1 :]
                col_mapping = _build_col_index_to_standard_name(header_row)
                if len(col_mapping) < 2 and data_rows:
                    col_mapping = _fallback_col_mapping(data_rows)
                score = _score_table_as_item_table(table)
                candidates.append((score, col_mapping, data_rows, table))
    except Exception:
        pass

    # 2) find_tables (lines) 로 추가 후보
    try:
        table_settings_lines = {
            "vertical_strategy": "lines",
            "horizontal_strategy": "lines",
            "snap_tolerance": 3,
            "join_tolerance": 3,
        }
        tables = page.find_tables(table_settings=table_settings_lines)
        if tables:
            for t in tables:
                table = t.extract()
                if not table or len(table) < 2:
                    continue
                header_idx = _find_header_row_index(table)
                header_row = table[header_idx]
                data_rows = table[header_idx + 1 :]
                col_mapping = _build_col_index_to_standard_name(header_row)
                if len(col_mapping) < 2 and data_rows:
                    col_mapping = _fallback_col_mapping(data_rows)
                score = _score_table_as_item_table(table)
                candidates.append((score, col_mapping, data_rows, table))
    except Exception:
        pass

    # 3) find_tables (text) 로 추가 후보
    try:
        table_settings_text = {
            "vertical_strategy": "text",
            "horizontal_strategy": "text",
            "snap_x_tolerance": 3,
            "snap_y_tolerance": 3,
        }
        tables = page.find_tables(table_settings=table_settings_text)
        if tables:
            for t in tables:
                table = t.extract()
                if not table or len(table) < 2:
                    continue
                header_idx = _find_header_row_index(table)
                header_row = table[header_idx]
                data_rows = table[header_idx + 1 :]
                col_mapping = _build_col_index_to_standard_name(header_row)
                if len(col_mapping) < 2 and data_rows:
                    col_mapping = _fallback_col_mapping(data_rows)
                score = _score_table_as_item_table(table)
                candidates.append((score, col_mapping, data_rows, table))
    except Exception:
        pass

    # 4) extract_table() 단일 테이블
    try:
        table = page.extract_table()
        if table and len(table) >= 2:
            header_idx = _find_header_row_index(table)
            header_row = table[header_idx]
            data_rows = table[header_idx + 1 :]
            col_mapping = _build_col_index_to_standard_name(header_row)
            if len(col_mapping) < 2 and data_rows:
                col_mapping = _fallback_col_mapping(data_rows)
            score = _score_table_as_item_table(table)
            candidates.append((score, col_mapping, data_rows, table))
    except Exception:
        pass

    # 품목 테이블 점수가 가장 높은 후보 선택 (최소 1개 이상 품목 키워드 또는 폴백으로 컬럼 있음)
    if not candidates:
        return None, []

    candidates.sort(key=lambda x: (-x[0], -len(x[2])))  # 점수 높은 순, 그다음 데이터 행 많은 순
    best = candidates[0]
    score, col_mapping, data_rows, _ = best
    # 점수 0이어도 폴백 매핑이 있으면 사용 (데이터 행에 실제 값이 있을 수 있음)
    return col_mapping, data_rows


def extract_commercial_invoice(pdf_path):
    """
    Commercial Invoice PDF에서 표 데이터를 전부 추출합니다.
    - 파일명에 'Invoice' 또는 'CI'가 포함된 파일 대상.
    - 페이지 상단에서 Invoice No, Invoice Date 추출 후, 선/텍스트 기반으로 테이블 추출.
    - 헤더 행 자동 감지 후 항목별(Order No, Product Code, Description, HTS No, Qty, Unit Price, Total) 매핑.
    - Description이 여러 행인 경우 \\n으로 합쳐서 한 행으로 처리.
    - 반환: list of dict (각 행이 하나의 레코드).
    """
    standard_columns = [
        "Invoice No",
        "Invoice Date",
        "Order No",
        "Product Code",
        "Description",
        "Gender",
        "Brand Description",
        "Product Category",
        "Quality",
        "HTS No",
        "Country of Origin",
        "Qty",
        "UOM",
        "Unit Price",
        "Total",
        "Manufacturer",
    ]
    records = []
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page_num, page in enumerate(pdf.pages):
                # 1. 페이지 상단에서 Invoice No, Invoice Date 추출
                invoice_no, invoice_date = extract_invoice_header(page)

                # 2. 여러 전략으로 테이블 추출 후 헤더/데이터 행 분리
                col_mapping, data_rows = _extract_invoice_table_from_page(page)
                if not col_mapping or not data_rows:
                    continue

                # 3. PDF 한 박스가 여러 행으로 나뉘어 추출된 경우 한 행으로 합침 (한 박스 = 한 셀)
                # Order No만 키로 사용: Order No가 비어 있으면 연속 행(같은 박스의 다음 줄)으로 간주
                key_col_indices = sorted(
                    [i for i, name in col_mapping.items() if name == "Order No"]
                )
                if not key_col_indices and col_mapping:
                    key_col_indices = [min(col_mapping.keys())]  # 첫 컬럼을 키로 fallback
                if key_col_indices:
                    data_rows = merge_continuation_rows(data_rows, key_col_indices)
                # Description만 여러 행인 경우 추가 합치기 (fallback)
                desc_col_index = None
                for idx, std_name in col_mapping.items():
                    if std_name == "Description":
                        desc_col_index = idx
                        break
                if desc_col_index is not None:
                    data_rows = merge_description_rows(data_rows, desc_col_index)

                # 4. col_mapping이 모든 데이터 행의 컬럼 수를 커버하도록 확장 (행마다 셀 수가 다를 수 있음)
                max_cols = max((len(list(r)) for r in data_rows), default=0)
                fallback_names = [
                    "Order No", "Product Code", "Description", "HTS No", "Qty", "Unit Price", "Total"
                ]
                for i in range(len(col_mapping), max_cols):
                    if i < len(fallback_names):
                        col_mapping[i] = fallback_names[i]

                # 5. 각 데이터 행을 표준 컬럼 dict로 변환 (col_mapping의 모든 인덱스에 대해 값 채움)
                # 원칙: PDF 한 박스(셀) = 엑셀 한 셀. 셀 내용을 줄 단위로 나누지 않고 그대로 한 셀에 넣음.
                page_records = []
                for row in data_rows:
                    row = list(row) if row else []
                    rec = {col: "" for col in standard_columns}
                    for i, std_name in col_mapping.items():
                        val = row[i] if i < len(row) else ""
                        if std_name in rec:
                            rec[std_name] = val if val is not None else ""
                            if isinstance(rec[std_name], str):
                                rec[std_name] = rec[std_name].strip()
                    # 페이지에서 추출한 Invoice No, Invoice Date로 보정
                    if invoice_no:
                        rec["Invoice No"] = invoice_no
                    if invoice_date:
                        rec["Invoice Date"] = invoice_date
                    # 박스 안 모든 문자·숫자 보존. 줄바꿈(\n)은 한 셀 내 여러 줄로 유지.
                    # HTS No: "HTS No.", "No." 등 헤더 유출 제거 (No.6402999000 → 6402999000). 항상 선행 제거 적용
                    if rec.get("HTS No") and isinstance(rec["HTS No"], str):
                        s = rec["HTS No"].replace("\r", "").strip()
                        s = re.sub(r"^(hts\s*no\.?|no\.?)\s*", "", s, flags=re.I)
                        if re.match(r"^[\d\n\s\.]+$", s):
                            rec["HTS No"] = s.replace("\n", "").replace(" ", "").strip()
                        else:
                            rec["HTS No"] = s.strip()
                    # Brand Description / Description / Product Category / Country of Origin: 헤더 유출 제거
                    # (DescriptionCrocs → Crocs). 접두사가 없을 때까지 반복 제거
                    for key in ("Brand Description", "Description", "Product Category", "Country of Origin"):
                        if rec.get(key) and isinstance(rec[key], str):
                            t = rec[key].strip()
                            prefixes = (
                                "Description", "Category", "Brand", "Product ", "Product category ",
                                "Quality ", "Gender ", "of Origin ", "of Origin",
                            )
                            while True:
                                changed = False
                                for prefix in prefixes:
                                    if t.lower().startswith(prefix.lower()) and len(t) > len(prefix):
                                        t = t[len(prefix):].strip()
                                        changed = True
                                        break
                                if not changed:
                                    break
                            rec[key] = _deduplicate_repeated_phrase(t)
                    # Manufacturer 등에도 중복 문구 제거 적용
                    for key in ("Manufacturer",):
                        if rec.get(key) and isinstance(rec[key], str):
                            rec[key] = _deduplicate_repeated_phrase(rec[key].strip())
                    # Product Code/Description 셀 경계 오분리 보정 (Classic→assic, Echo→o 등)
                    _fix_product_code_description_split(rec)
                    # Unit Price 누락 시 Total/Qty로 계산; Total·Qty 있으면 Unit Price 역산
                    try:
                        q = rec.get("Qty")
                        t = rec.get("Total")
                        if (q is not None and t is not None) and (not rec.get("Unit Price") or not str(rec.get("Unit Price")).strip()):
                            qn = pd.to_numeric(str(q).replace(",", "").replace(" ", ""), errors="coerce")
                            tn = pd.to_numeric(str(t).replace(",", "").replace(" ", ""), errors="coerce")
                            if qn and tn and float(qn) != 0:
                                rec["Unit Price"] = float(tn) / float(qn)
                    except Exception:
                        pass
                    # 헤더/상단 블록 행 제외
                    order_no_str = str(rec.get("Order No") or "").strip().lower()
                    if order_no_str and "invoice date" in order_no_str:
                        continue
                    if order_no_str and any(x in order_no_str for x in ("net total", "value added", "grand total", "total amount")):
                        continue
                    has_data = (
                        rec.get("Invoice No")
                        or rec.get("Order No")
                        or rec.get("Product Code")
                        or rec.get("Description")
                        or rec.get("Qty")
                        or rec.get("Unit Price")
                        or rec.get("Total")
                    )
                    if has_data:
                        page_records.append(rec)

                # 6. 병합 셀 보정: Order No, Product Code 등이 위쪽 행에만 있고 아래 행은 빈 경우, 이전 행 값으로 채움
                _forward_fill_merged_cells(page_records)
                # 7. 서브헤더 행(No. 등): 다음 행의 Order No/Product Code로 back-fill (테이블 상단에 헤더만 있고 첫 품목이 다음 행인 경우)
                _back_fill_merged_cells(page_records)
                # 8. 동일 Order No로 두 행으로 나뉜 경우 병합 (첫 행에 Order No·Product Code만, 다음 행에 Description·Qty 등)
                _merge_same_order_no_rows(page_records)
                records.extend(page_records)
    except Exception as e:
        raise RuntimeError(f"인보이스 PDF 처리 중 오류 발생: {pdf_path}") from e

    return records


# =============================================================================
# 유형 2: Certificate of Origin 추출 (PaddleOCR)
# =============================================================================


def preprocess_image_for_ocr(pil_image):
    """
    붉은색 관인(Stamp) 등 노이즈를 줄이기 위한 전처리.
    Grayscale 변환 후 이진화(Adaptive Threshold) 적용.
    """
    img = np.array(pil_image)
    if len(img.shape) == 3:
        gray = cv2.cvtColor(img, cv2.COLOR_RGB2GRAY)
    else:
        gray = img.copy()

    # 적응형 이진화: 붉은색 스탬프가 회색으로 남아도 경계가 선명해짐
    binary = cv2.adaptiveThreshold(
        gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 11, 2
    )
    return binary


def extract_text_from_co_image_tesseract(pil_image):
    """
    Tesseract OCR로 이미지에서 텍스트 추출 (PaddleOCR 대안).
    (box, text) 리스트 반환. parse_co_fields와 호환되도록 텍스트만 사용.
    """
    img_np = preprocess_image_for_ocr(pil_image)
    # pytesseract는 numpy array 또는 PIL 지원
    text = pytesseract.image_to_string(img_np, lang="eng")
    if not text or not text.strip():
        return []
    # 한 줄씩 (box는 None) 리스트로 반환하여 parse_co_fields와 호환
    return [(None, line.strip()) for line in text.splitlines() if line.strip()]


def extract_text_from_co_image(pil_image, ocr_engine):
    """
    이미지 한 장에서 OCR로 텍스트 추출.
    ocr_engine이 있으면 PaddleOCR, 없으면 Tesseract 사용. (좌표, 텍스트) 리스트 반환.
    """
    img_preprocessed = preprocess_image_for_ocr(pil_image)

    if ocr_engine is not None:
        # PaddleOCR 경로
        result = ocr_engine.ocr(img_preprocessed, cls=True)
        if not result or not result[0]:
            return []
        lines = []
        for line in result[0]:
            if line and len(line) >= 2:
                box, (text, _) = line[0], line[1]
                if text and str(text).strip():
                    lines.append((box, str(text).strip()))
        return lines

    # Tesseract 경로 (Paddle 미사용 시)
    if _HAS_TESSERACT:
        return extract_text_from_co_image_tesseract(pil_image)
    return []


def parse_co_fields(text_lines):
    """
    CO OCR 결과 텍스트에서 다음 항목을 정규식으로 추출:
    - Reference No (우측 상단, B26MA4... 형태)
    - Origin Criterion (8번 항목, "B" 45% 등)
    - Invoice No (10번 항목)
    - Tariff Item Number (5번 항목, HS Code)
    """
    full_text = " ".join(t for (_, t) in text_lines)

    reference_no = None
    origin_criterion = None
    invoice_no = None
    tariff_item_number = None

    # Reference No: 영문+숫자 조합 (우측 상단에 자주 등장)
    ref_match = re.search(r"\b([A-Z][0-9]{2}[A-Z]{2}[A-Z0-9]{2,})\b", full_text)
    if ref_match:
        reference_no = ref_match.group(1)

    # 번호 항목 패턴: "1.", "2." ... "10." 등으로 시작하는 블록
    # 5. Tariff Item Number (HS Code)
    tariff_match = re.search(
        r"5\.\s*(?:Tariff\s*Item\s*Number|HS\s*Code)?\s*:?\s*([0-9]{4}\.?[0-9]{2}\.?[0-9]{2})",
        full_text,
        re.I,
    )
    if not tariff_match:
        tariff_match = re.search(r"\b([0-9]{4}\.[0-9]{2}\.[0-9]{2})\b", full_text)
    if tariff_match:
        tariff_item_number = tariff_match.group(1).replace(" ", "")

    # 8. Origin Criterion ("B" 45% 등)
    origin_match = re.search(
        r"8\.\s*(?:Origin\s*Criterion)?\s*:?\s*([A-Z])\s*(\d+\s*%?)?",
        full_text,
        re.I,
    )
    if origin_match:
        origin_criterion = origin_match.group(1)
        if origin_match.group(2):
            origin_criterion = origin_criterion + " " + origin_match.group(2).strip()

    # 10. Invoice No (인보이스와 매칭 키)
    inv_match = re.search(
        r"10\.\s*(?:Invoice\s*No\.?)?\s*:?\s*([A-Za-z0-9\-]+)",
        full_text,
        re.I,
    )
    if inv_match:
        invoice_no = inv_match.group(1).strip()
    if not invoice_no:
        inv_match = re.search(r"Invoice\s*No\.?\s*:?\s*([A-Za-z0-9\-]+)", full_text, re.I)
        if inv_match:
            invoice_no = inv_match.group(1).strip()

    return {
        "Reference No": reference_no,
        "Origin Criterion": origin_criterion,
        "Invoice No": invoice_no,
        "Tariff Item Number": tariff_item_number,
    }


def extract_certificate_of_origin(pdf_path):
    """
    Certificate of Origin PDF를 이미지로 변환 후 OCR로 텍스트 추출하고,
    Reference No, Origin Criterion, Invoice No, Tariff Item Number를 파싱합니다.
    - PaddleOCR 설치 시 Paddle 사용, 미설치 시 Tesseract 사용 (pytesseract + Tesseract 엔진 필요).
    - 파일명에 'Certificate' 또는 'Origin'이 포함된 파일 대상.
    - 반환: list of dict (CO는 보통 1건이므로 1개 요소인 리스트).
    """
    if not _HAS_PADDLE and not _HAS_TESSERACT:
        raise RuntimeError(
            "원산지 증명서(CO) OCR을 위해 paddleocr 또는 pytesseract 중 하나가 필요합니다. "
            "pip install pytesseract 후, Windows에서는 Tesseract 엔진을 설치하고 PATH에 추가하세요."
        )

    try:
        images = pdf2image.convert_from_path(pdf_path, dpi=200)
    except Exception as e:
        raise RuntimeError(f"PDF를 이미지로 변환 실패: {pdf_path}") from e

    ocr_engine = get_ocr_engine()
    all_text_lines = []

    for pil_img in images:
        lines = extract_text_from_co_image(pil_img, ocr_engine)
        all_text_lines.extend(lines)

    if not all_text_lines:
        return []

    fields = parse_co_fields(all_text_lines)
    return [fields]


# =============================================================================
# 데이터 통합 및 엑셀 저장
# =============================================================================

# Commercial Invoice 엑셀 형식: 참조 파일(CommercialInvoice_... CI 1.xlsx)과 동일한 레이아웃
# 상단: Invoice No, Invoice Date (1~2행) / 3~14행: 빈 행 또는 제목·Seller/Buyer 등(비워둠)
# 15행: 컬럼 헤더, 16행~: 데이터
CI_EXCEL_HEADER_ROW = 15  # 1-based; 컬럼 헤더가 있는 행
CI_EXCEL_DATA_START_ROW = 16  # 데이터 시작 행
# 대상 컬럼 순서 (참조 Commercial Invoice CI 1.xlsx)
CI_COLUMN_HEADERS = [
    "Order No.",
    "Product Code",
    "Description",
    "Gender",
    "Brand Description",
    "Product Category",
    "Quality",
    "HTS No.",
    "Country of Origin",
    "Qty",
    "UOM",
    "Unit Price",
    None,  # 빈 열
    "Value Add Price",
    "Gross Unit Price",
    "Total",
    "Manufacturer",
]


def write_excel_commercial_invoice_format(df_merged, output_path, invoice_no=None, invoice_date=None):
    """
    추출된 데이터를 Commercial Invoice 엑셀 형식(참조: CommercialInvoice_... CI 1.xlsx)으로 저장합니다.
    - 1행: Invoice No: | 값
    - 2행: Invoice Date: | 값
    - 3~14행: 빈 행 (추후 Seller/Buyer 등 채울 수 있음)
    - 15행: Order No., Product Code, Description, ... (테이블 헤더)
    - 16행~: 데이터 행 (우리 추출 컬럼을 대상 컬럼 순서에 매핑)
    """
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 상단 블록: Invoice No, Invoice Date
    first_inv = invoice_no or (df_merged["Invoice No"].iloc[0] if not df_merged.empty and "Invoice No" in df_merged.columns else "")
    first_date = invoice_date or (df_merged["Invoice Date"].iloc[0] if not df_merged.empty and "Invoice Date" in df_merged.columns else "")
    ws.cell(row=1, column=1, value="Invoice No:")
    ws.cell(row=1, column=2, value=first_inv)
    ws.cell(row=2, column=1, value="Invoice Date:")
    ws.cell(row=2, column=2, value=first_date)
    # 4행: Commercial Invoice
    ws.cell(row=4, column=1, value="Commercial Invoice")
    # 6행: Page 1 of 1
    ws.cell(row=6, column=1, value="Page: 1 of 1")

    # 15행: 테이블 헤더
    for col_idx, header in enumerate(CI_COLUMN_HEADERS, start=1):
        if header:
            ws.cell(row=CI_EXCEL_HEADER_ROW, column=col_idx, value=header)

    # 우리 DataFrame 컬럼 → 대상 컬럼 인덱스 매핑
    # Order No. <- Order No, Product Code <- Product Code, Description <- Description, ...
    our_to_ci_index = {
        "Order No": 1,
        "Product Code": 2,
        "Description": 3,
        "Gender": 4,
        "Brand Description": 5,
        "Product Category": 6,
        "Quality": 7,
        "HTS No": 8,
        "Country of Origin": 9,
        "Qty": 10,
        "UOM": 11,
        "Unit Price": 12,
        "Total": 16,
        "Manufacturer": 17,
    }
    # Country of Origin: CO에서 Tariff/Origin 있으면 사용
    if "Tariff Item Number" in df_merged.columns:
        our_to_ci_index["Tariff Item Number"] = 8  # HTS와 같은 열에 넣지 말고, HTS No는 우리 HTS No
    # Gross Unit Price: Unit Price와 동일 값
    ci_col_count = len(CI_COLUMN_HEADERS)

    for data_row_num, (r_idx, df_row) in enumerate(df_merged.iterrows()):
        excel_row = CI_EXCEL_DATA_START_ROW + data_row_num
        for col_name, ci_col in our_to_ci_index.items():
            if col_name not in df_merged.columns:
                continue
            val = df_row.get(col_name)
            if pd.isna(val):
                val = ""
            # Order No, Product Code: 숫자여도 문자열로 저장 (과학적 표기 방지)
            if col_name in ("Order No", "Product Code") and isinstance(val, (int, float)):
                val = str(int(val)) if val == int(val) else str(val)
            if isinstance(val, float) and col_name in ("Qty", "Unit Price", "Total"):
                try:
                    if col_name == "Total" and abs(val) >= 1e6:
                        ws.cell(row=excel_row, column=ci_col, value=f"{val:,.4f}")
                    else:
                        ws.cell(row=excel_row, column=ci_col, value=val)
                except Exception:
                    ws.cell(row=excel_row, column=ci_col, value=val)
            else:
                ws.cell(row=excel_row, column=ci_col, value=val)
        # Gross Unit Price (15열) = Unit Price와 동일
        unit_price_val = df_row.get("Unit Price")
        if unit_price_val is not None and not (isinstance(unit_price_val, float) and pd.isna(unit_price_val)):
            ws.cell(row=excel_row, column=15, value=unit_price_val)

    wb.save(output_path)


def numeric_columns_clean(df, numeric_cols):
    """지정한 컬럼을 pd.to_numeric으로 float 변환 (errors='coerce')."""
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(
                df[col].astype(str).str.replace(",", "").str.replace(" ", ""),
                errors="coerce",
            )
    return df


def collect_pdf_paths(directory):
    """디렉터리에서 Invoice/CI, Certificate/Origin PDF 경로를 분류하여 반환."""
    directory = Path(directory)
    if not directory.is_dir():
        return [], []

    invoice_files = []
    co_files = []
    for path in directory.glob("**/*.pdf"):
        name = path.name.lower()
        if "invoice" in name or "ci" in name:
            invoice_files.append(path)
        if "certificate" in name or "origin" in name:
            co_files.append(path)

    return invoice_files, co_files


def run_extraction_and_merge(source_directory, output_excel_path="Customs_Integrated_Data.xlsx"):
    """
    source_directory 내 PDF를 스캔하여 인보이스/CO를 추출하고,
    Invoice No 기준으로 outer merge 후 엑셀 저장합니다.
    """
    invoice_paths, co_paths = collect_pdf_paths(source_directory)

    # 인보이스 추출
    invoice_records = []
    for path in invoice_paths:
        try:
            recs = extract_commercial_invoice(str(path))
            invoice_records.extend(recs)
        except Exception as e:
            _safe_print(f"[경고] 인보이스 처리 실패, 다음 파일로 진행: {path}\n오류: {e}")

    # CO 추출
    co_records = []
    for path in co_paths:
        try:
            recs = extract_certificate_of_origin(str(path))
            co_records.extend(recs)
        except Exception as e:
            _safe_print(f"[경고] 원산지 증명서 처리 실패, 다음 파일로 진행: {path}\n오류: {e}")

    if not invoice_records and not co_records:
        print("추출된 데이터가 없습니다. PDF 경로를 확인하세요.")
        return 0

    df_inv = pd.DataFrame(invoice_records) if invoice_records else pd.DataFrame()
    df_co = pd.DataFrame(co_records) if co_records else pd.DataFrame()

    # 숫자 컬럼 변환
    numeric_cols = ["Qty", "Unit Price", "Total"]
    if not df_inv.empty:
        df_inv = numeric_columns_clean(df_inv, numeric_cols)

    # Invoice No 기준 outer merge (여러 아이템이 있어도 행 단위로 결합)
    if df_inv.empty:
        df_merged = df_co
    elif df_co.empty:
        df_merged = df_inv
    else:
        # CO는 Invoice No당 보통 1행이므로, 같은 Invoice No를 가진 인보이스 행마다 CO 정보가 반복됨
        df_merged = pd.merge(
            df_inv,
            df_co,
            on="Invoice No",
            how="outer",
            suffixes=("", "_co"),
        )
        # Invoice No가 없는 CO 행도 포함되도록 이미 outer로 처리됨

    # 중복 컬럼 정리 (merge 시 _co 접미사가 붙은 경우)
    cols = [c for c in df_merged.columns if not c.endswith("_co")]
    df_merged = df_merged[cols]

    # Commercial Invoice 형식(참조: CommercialInvoice_... CI 1.xlsx)으로 저장
    write_excel_commercial_invoice_format(df_merged, output_excel_path)
    print(f"저장 완료: {output_excel_path} (총 {len(df_merged)}행, Commercial Invoice 형식)")
    return len(df_merged)


# =============================================================================
# 메인 진입점
# =============================================================================

if __name__ == "__main__":
    # PDF가 있는 폴더 경로 (현재 스크립트 위치 또는 지정 경로)
    script_dir = Path(__file__).resolve().parent
    pdf_folder = script_dir / "pdf_files"  # pdf_files 폴더에 CI, CO PDF를 넣어두면 됨

    # pdf_files가 없으면 현재 디렉터리에서 PDF 검색
    if not pdf_folder.is_dir():
        pdf_folder = script_dir

    run_extraction_and_merge(str(pdf_folder), "Customs_Integrated_Data.xlsx")
