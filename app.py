# -*- coding: utf-8 -*-
"""
해외실증지원사업 모니터링 시스템
- 지원기업: 로그인 후 본인 기업만 조회
- 한국환경산업기술원/관세법인 드림: 모든 기업 정보 조회
"""
from __future__ import annotations

from datetime import datetime as dt_util
from pathlib import Path
from typing import List, Optional

import matplotlib.pyplot as plt
import pandas as pd
import streamlit as st
from sqlalchemy import or_
from sqlalchemy.orm import Session

from openpyxl import load_workbook

from auth import authenticate, get_company_for_user, hash_password
from database import get_session, init_db
from models import (
    CloseSupport,
    Company,
    Consultant,
    DeliveryTimelineEvent,
    DocumentItem,
    DocChecklistStatus,
    ExecutionPlan,
    ExportStep,
    HSCodeReport,
    PendingProfileUpdate,
    PreAnalysis,
    PreDiagnosisAttachment,
    PreDiagnosisChecklistResponse,
    PreDiagnosisChecklistTemplate,
    HistoryEntry,
    ResultReport,
    TradeEnvironmentReport,
    User,
    UserRole,
)

BASE_DIR = Path(__file__).resolve().parent
UPLOAD_DIR = BASE_DIR / "uploads"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
PREDIAG_EXCEL_PATH = BASE_DIR / "사전진단체크리스트.xlsx"


def _cell_str(val) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    return s if s else ""


def load_prediag_excel_structure(excel_path: Optional[Path] = None):
    """
    엑셀 파일의 병합(merged_cells) 구조를 읽어 구분·항목·내용·관련서류 등 구조를 반환.
    반환: list of dict with keys: section, item, content_placeholder, related_doc, guide_notes, section_rowspan
    section_rowspan: 구분 열의 rowspan (0이면 해당 행은 병합된 셀의 연속이므로 구분 셀 미표시)
    파일이 없거나 읽기 실패 시 None 반환.
    """
    path = excel_path or PREDIAG_EXCEL_PATH
    if not path.exists():
        return None
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        # 병합 영역: (row, col) -> (min_row, min_col, max_row, max_col)
        merged = {}
        for mr in ws.merged_cells.ranges:
            for r in range(mr.min_row, mr.max_row + 1):
                for c in range(mr.min_col, mr.max_col + 1):
                    merged[(r, c)] = (mr.min_row, mr.min_col, mr.max_row, mr.max_col)
        rows = []
        current_section = ""
        # 엑셀 1-based: B=구분(2), C=항목(3), D=내용(4), E=관련서류(5), F,G,H=수취·요청·없음(6,7,8), I=비고(9), J=Comment(10)
        for row_idx in range(4, ws.max_row + 1):
            val_b = ws.cell(row=row_idx, column=2).value
            val_c = ws.cell(row=row_idx, column=3).value
            val_d = ws.cell(row=row_idx, column=4).value
            val_e = ws.cell(row=row_idx, column=5).value
            val_i = ws.cell(row=row_idx, column=9).value
            val_j = ws.cell(row=row_idx, column=10).value
            # 헤더 행 또는 항목이 비어 있으면 스킵 (데이터 행만 수집)
            item_str = _cell_str(val_c)
            if not item_str or item_str == "항목":
                continue
            # 구분: 병합이면 상단 셀 값 사용
            merge_info = merged.get((row_idx, 2))
            if merge_info:
                min_r, min_c, max_r, max_c = merge_info
                section_val = ws.cell(row=min_r, column=2).value
                section_str = _cell_str(section_val)
                section_rowspan = (max_r - min_r + 1) if row_idx == min_r else 0
            else:
                section_str = _cell_str(val_b) or current_section
                section_rowspan = 1
            if section_str:
                current_section = section_str
            rows.append({
                "section": section_str or current_section,
                "item": item_str,
                "content_placeholder": _cell_str(val_d),
                "related_doc": _cell_str(val_e),
                "guide_notes": _cell_str(val_i),
                "comment_placeholder": _cell_str(val_j),
                "section_rowspan": section_rowspan,
            })
        wb.close()
        return rows
    except Exception:
        return None


# 메뉴 키
MENU_DASHBOARD = "대시보드"
MENU_PRE_DIAG = "사전진단"
MENU_HS_CODE = "HS code관리"
MENU_EXEC_PLAN = "수행계획서"
MENU_DOC_CHECKLIST = "서류관리"
MENU_DELIVERY_TIMELINE = "배송관리"
MENU_STAGE_BOARD = "전체 진행단계별 보드"
MENU_RESULT_REPORT = "결과보고서"
MENU_HISTORY = "히스토리 관리"

MENU_MY_PROFILE = "개인정보 설정"
MENU_ADMIN = "관리자 설정"
MENU_COMPANY_REGISTER = "지원기업 등록"
MENU_CONSULTANT_REGISTER = "관세사 등록"
MENU_COMPANY_DETAIL = "기업 상세"
MENU_ASSIGN_CONSULTANT = "담당 관세사 배정"

# 레벨1 메뉴
MENU_LEVEL1_DASHBOARD = "대시보드"
MENU_LEVEL1_PROGRESS = "진행관리"
MENU_LEVEL1_MISC = "기타관리"

# 메뉴 아이콘(픽토그램)
MENU_ICONS = {
    MENU_DASHBOARD: "📊",
    MENU_PRE_DIAG: "🔍",
    MENU_HS_CODE: "🏷️",
    MENU_EXEC_PLAN: "📄",
    MENU_DOC_CHECKLIST: "✅",
    MENU_DELIVERY_TIMELINE: "📅",
    MENU_STAGE_BOARD: "📋",
    MENU_RESULT_REPORT: "🏁",
    MENU_HISTORY: "🕒",
    MENU_MY_PROFILE: "👤",
    MENU_ADMIN: "⚙️",
    MENU_COMPANY_REGISTER: "🏢",
    MENU_CONSULTANT_REGISTER: "👔",
    MENU_COMPANY_DETAIL: "🏷️",
    MENU_ASSIGN_CONSULTANT: "🤝",
}


def company_progress_status(session: Session, company: Company) -> dict:
    """요구사항 기반 진행상태 산정(테이블/대시보드 공통 사용)."""
    # 사전진단 완료: 체크리스트 응답 1개라도 있으면 완료로 판단
    prediag_done = (
        session.query(PreDiagnosisChecklistResponse)
        .filter(PreDiagnosisChecklistResponse.company_id == company.id)
        .count()
        > 0
    )
    hs_done = (
        session.query(HSCodeReport)
        .filter(HSCodeReport.company_id == company.id)
        .count()
        > 0
    )
    exec_done = (
        session.query(ExecutionPlan)
        .filter(ExecutionPlan.company_id == company.id)
        .count()
        > 0
    )
    docs_done = bool(company.doc_checklist_completed)
    result_done = (
        session.query(ResultReport)
        .filter(ResultReport.company_id == company.id)
        .count()
        > 0
    )

    # 현재 진행단계(요약 표시용)
    if result_done:
        current = "배송완료"
    elif not prediag_done:
        current = "사전진단"
    elif not hs_done:
        current = "HS code확정"
    elif not exec_done:
        current = "수행계획서"
    elif not docs_done:
        current = "서류준비"
    else:
        current = "배송준비"

    return {
        "prediag_done": prediag_done,
        "hs_done": hs_done,
        "exec_done": exec_done,
        "docs_done": docs_done,
        "result_done": result_done,
        "current": current,
    }


def compute_phase_statuses(session: Session, company: Company, ps: dict) -> dict:
    """대시보드용 단계별 상태(사전진단/품목분류/수행계획/서류구비/물류준비/배송/완료)."""
    # 공통 상태 맵
    def status_label(done: bool) -> str:
        return "완료" if done else "미결"

    # 사전진단 / 품목분류 / 수행계획
    prediag = status_label(ps["prediag_done"])
    hs = status_label(ps["hs_done"])
    exec_plan = status_label(ps["exec_done"])

    # 서류구비: 체크리스트 항목 중 체크 수에 따라 미결/진행 중/완료
    total_docs = (
        session.query(DocumentItem)
        .filter(or_(DocumentItem.company_id.is_(None), DocumentItem.company_id == company.id))
        .count()
    )
    checked_docs = (
        session.query(DocChecklistStatus)
        .filter(DocChecklistStatus.company_id == company.id, DocChecklistStatus.is_checked == 1)
        .count()
    )
    if checked_docs == 0:
        docs = "미결"
    elif total_docs and checked_docs == total_docs:
        docs = "완료"
    else:
        docs = "진행 중"

    # 물류준비: 배송일정 타임라인의 선적준비 단계 완료 여부
    shipping_event = (
        session.query(DeliveryTimelineEvent)
        .filter(
            DeliveryTimelineEvent.company_id == company.id,
            DeliveryTimelineEvent.stage == ExportStep.PRE_SHIPMENT,
        )
        .first()
    )
    logistics = "완료" if (shipping_event and shipping_event.actual_completed_at) else "미결"

    # 배송: 현재 배송 단계(타임라인 기준) 텍스트만 표시
    evs = (
        session.query(DeliveryTimelineEvent)
        .filter(DeliveryTimelineEvent.company_id == company.id)
        .all()
    )
    by_stage = {e.stage: e for e in evs}
    shipping_stage = "-"
    for step in ExportStep:
        ev = by_stage.get(step)
        if ev and (ev.actual_completed_at or ev.planned_date):
            shipping_stage = step.label
    shipping = shipping_stage

    # 완료: 결과보고서 등록 여부 기준 (추후 KEITI 승인 단계 추가 가능)
    if not ps["result_done"]:
        done = "미결"
    else:
        done = "진행 중"

    return {
        "사전진단": prediag,
        "품목분류": hs,
        "수행계획": exec_plan,
        "서류구비": docs,
        "물류준비": logistics,
        "배송": shipping,
        "완료": done,
    }


def log_history(session: Session, company_id: int, process: str, status: str) -> None:
    """공통 히스토리 기록 헬퍼. actor는 현재 로그인 아이디 사용."""
    actor = st.session_state.get("logged_in_login_id")
    entry = HistoryEntry(
        company_id=company_id,
        process=process,
        status=status,
        actor=actor,
    )
    session.add(entry)
    session.commit()


def save_uploaded_file(company_id: int, phase: str, file) -> Optional[str]:
    if file is None:
        return None
    phase_dir = UPLOAD_DIR / f"company_{company_id}" / phase
    phase_dir.mkdir(parents=True, exist_ok=True)
    file_path = phase_dir / file.name
    with open(file_path, "wb") as f:
        f.write(file.getbuffer())
    return str(file_path.relative_to(BASE_DIR))


def get_visible_companies(
    session: Session, role: UserRole, company_id: Optional[int]
) -> List[Company]:
    """권한에 따라 조회 가능한 기업 목록. 지원기업은 본인 기업만."""
    if role in (UserRole.ADMIN, UserRole.DREAM, UserRole.KEITI):
        return session.query(Company).order_by(Company.name).all()
    if role == UserRole.COMPANY and company_id:
        c = session.query(Company).filter(Company.id == company_id).first()
        return [c] if c else []
    return []


def get_selected_company(
    session: Session, companies: List[Company], company_id: Optional[int], key: str
) -> Optional[Company]:
    """지원기업이면 단일 기업, 아니면 셀렉트박스로 선택."""
    if not companies:
        return None
    if len(companies) == 1:
        return companies[0]
    names = [c.name for c in companies]
    default_idx = 0
    if company_id:
        for i, c in enumerate(companies):
            if c.id == company_id:
                default_idx = i
                break
    selected_name = st.selectbox("기업 선택", names, index=default_idx, key=key)
    return next(c for c in companies if c.name == selected_name)


def render_login_or_role(session: Session) -> bool:
    """통합 로그인: 모든 역할(관리자·관세법인·KEITI·지원기업) 아이디/비밀번호 로그인.

    브라우저 뒤로가기 등으로 세션이 새로 열려도, URL 쿼리파라미터에 저장된 login_id로
    다시 사용자 정보를 복원해 로그인 상태를 유지한다.
    """
    # 1) 쿼리파라미터 기반 로그인 복원 (세션이 새로 열린 경우)
    if not st.session_state.get("user_id") and "user" in st.query_params:
        login_from_qs = st.query_params.get("user")
        if isinstance(login_from_qs, list):
            login_from_qs = login_from_qs[0] if login_from_qs else None
        if login_from_qs:
            u = session.query(User).filter(User.login_id == login_from_qs).first()
            if u and u.is_approved == 1:
                st.session_state["user_id"] = u.id
                st.session_state["role"] = u.role
                st.session_state["company_id"] = u.company_id
                st.session_state["logged_in_login_id"] = u.login_id

    # 2) 로그아웃 처리 (세션 + 쿼리파라미터 초기화)
    if st.sidebar.button("로그아웃", key="logout"):
        for key in ("user_id", "role", "company_id", "logged_in_login_id"):
            st.session_state.pop(key, None)
        for key in list(st.query_params):
            del st.query_params[key]  # 쿼리파라미터 제거
        st.rerun()

    # 3) 이미 로그인된 상태
    if st.session_state.get("user_id"):
        st.sidebar.success(f"로그인: {st.session_state.get('logged_in_login_id', '')}")
        return True

    # 4) 로그인 폼
    st.sidebar.markdown("### 로그인")
    login_id = st.sidebar.text_input("아이디", key="login_id_input")
    password = st.sidebar.text_input("비밀번호", type="password", key="password_input")
    if st.sidebar.button("로그인", key="login_btn"):
        user = authenticate(session, login_id, password)
        if user:
            st.session_state["user_id"] = user.id
            st.session_state["role"] = user.role
            st.session_state["company_id"] = user.company_id
            st.session_state["logged_in_login_id"] = user.login_id
            # 로그인 정보를 쿼리파라미터에 저장 (브라우저 뒤로가기/새 세션에서 복원용)
            st.query_params["user"] = user.login_id
            st.sidebar.success("로그인 성공")
            st.rerun()
        else:
            st.sidebar.error("아이디/비밀번호를 확인하세요. (승인 대기 중이면 로그인 불가)")
    return False


def set_nav(page: str, company_id: Optional[int] = None) -> None:
    """현재 메뉴/선택 기업을 URL에 저장해 브라우저 뒤로가기에 대응."""
    user_login_id = st.session_state.get("logged_in_login_id")
    if user_login_id:
        st.query_params["user"] = user_login_id
    st.query_params["page"] = page
    if company_id is not None:
        st.query_params["company"] = str(company_id)
        st.session_state["selected_company_id"] = company_id
    st.session_state["menu"] = page


def page_dashboard(session: Session, companies: List[Company]):
    st.markdown('<p class="section-title">📊 대시보드 — 기업별 전체 진행상황</p>', unsafe_allow_html=True)
    if not companies:
        st.info("조회 가능한 기업이 없습니다.")
        return

    # 관세법인 드림 요구사항: KPI는 '지원기업 등록 + 관세사 배정된 기업' 기준으로 산정
    assigned_companies = [c for c in companies if c.assigned_consultant]
    total = len(assigned_companies)
    not_started = 0
    completed = 0
    rows = []
    for c in assigned_companies:
        ps = company_progress_status(session, c)
        # 미진행: 사전진단 체크리스트 + 수행계획서 둘 다 미등록
        if (not ps["prediag_done"]) and (not ps["exec_done"]):
            not_started += 1
        if ps["result_done"]:
            completed += 1
        rows.append({
            "업체명": c.name,
            "담당 관세사": c.assigned_consultant or "",
            "현재 진행단계": ps["current"],
        })
    in_progress = total - not_started
    df = pd.DataFrame(rows) if rows else pd.DataFrame(columns=["업체명", "담당 관세사", "현재 진행단계"])

    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("총 지원기업", f"{total}개", "")
    with col2:
        st.metric("진행 중", f"{in_progress}개", "")
    with col3:
        st.metric("현지 도착(완료)", f"{completed}개", "")
    with col4:
        st.metric("미진행", f"{not_started}개", "")

    st.markdown("---")
    st.markdown("**지원기업 리스트 (기업명 클릭 → 상세페이지)**")
    if not assigned_companies:
        st.info("관세사가 배정된 지원기업이 없습니다. '담당 관세사 배정'에서 먼저 배정하세요.")
        return

    # 단계별 상태용 배지 렌더러
    def pill(text: str) -> str:
        status = text
        cls = "status-none"
        if status == "미결":
            cls = "status-pending"
        elif status == "완료":
            cls = "status-done"
        elif status == "진행 중":
            cls = "status-progress"
        return f'<span class="status-pill {cls}">{text}</span>'

    # 헤더 행
    h_cols = st.columns([2.2, 1.6, 1, 1, 1, 1, 1, 1, 1])
    headers = [
        "지원기업",
        "담당 관세사",
        "사전진단",
        "품목분류",
        "수행계획",
        "서류구비",
        "물류준비",
        "배송",
        "완료",
    ]
    for col, label in zip(h_cols, headers):
        col.markdown(f"**{label}**")

    # 각 기업 행
    for c in assigned_companies:
        ps = company_progress_status(session, c)
        phases = compute_phase_statuses(session, c, ps)
        cols_row = st.columns([2.2, 1.6, 1, 1, 1, 1, 1, 1, 1])
        with cols_row[0]:
            if st.button(c.name, key=f"goto_company_{c.id}"):
                set_nav(MENU_COMPANY_DETAIL, company_id=c.id)
                st.rerun()
        with cols_row[1]:
            st.write(c.assigned_consultant or "")
        with cols_row[2]:
            st.markdown(pill(phases["사전진단"]), unsafe_allow_html=True)
        with cols_row[3]:
            st.markdown(pill(phases["품목분류"]), unsafe_allow_html=True)
        with cols_row[4]:
            st.markdown(pill(phases["수행계획"]), unsafe_allow_html=True)
        with cols_row[5]:
            st.markdown(pill(phases["서류구비"]), unsafe_allow_html=True)
        with cols_row[6]:
            st.markdown(pill(phases["물류준비"]), unsafe_allow_html=True)
        with cols_row[7]:
            # 배송 단계명은 상태색 대신 회색 배지로만 표시
            st.markdown(f'<span class="status-pill status-none">{phases["배송"]}</span>', unsafe_allow_html=True)
        with cols_row[8]:
            st.markdown(pill(phases["완료"]), unsafe_allow_html=True)

    st.markdown("---")
    st.markdown("**전체 기업 진행상황 테이블**")
    st.dataframe(df, use_container_width=True, hide_index=True)


def page_pre_diagnosis(session: Session, company: Optional[Company], role: UserRole):
    st.subheader("사전진단 - 사전진단 결과 확인")
    if not company:
        return

    # ---- 사전진단 체크리스트 (엑셀 병합 구조 반영, HTML 테이블) ----
    templates = (
        session.query(PreDiagnosisChecklistTemplate)
        .order_by(PreDiagnosisChecklistTemplate.sort_order)
        .all()
    )
    responses_by_tid = {
        r.template_id: r
        for r in session.query(PreDiagnosisChecklistResponse)
        .filter(PreDiagnosisChecklistResponse.company_id == company.id)
        .all()
    }

    excel_rows = load_prediag_excel_structure()
    # display_rows: template_id, section, item, related_doc, guide_notes, section_rowspan, content, received, requested, none_status, comment
    def _section_rowspan_from_templates():
        out = []
        for i, t in enumerate(templates):
            if i == 0 or t.section != templates[i - 1].section:
                span = 1
                for j in range(i + 1, len(templates)):
                    if templates[j].section == t.section:
                        span += 1
                    else:
                        break
            else:
                span = 0
            out.append(span)
        return out

    fallback_rowspans = _section_rowspan_from_templates()
    display_rows = []
    for i, t in enumerate(templates):
        r = responses_by_tid.get(t.id)
        if excel_rows and i < len(excel_rows):
            ex = excel_rows[i]
            section_rowspan = ex.get("section_rowspan", 1)
            section = ex.get("section") or t.section
            item = ex.get("item") or t.item
            related_doc = ex.get("related_doc") or (t.related_doc or "")
            guide_notes = ex.get("guide_notes") or (t.guide_notes or "")
        else:
            section_rowspan = fallback_rowspans[i]
            section = t.section
            item = t.item
            related_doc = t.related_doc or ""
            guide_notes = (t.guide_notes or "")[:200]
        display_rows.append({
            "template_id": t.id,
            "section": section,
            "item": item,
            "related_doc": related_doc,
            "guide_notes": guide_notes,
            "section_rowspan": section_rowspan,
            "content": (r.content or "") if r else "",
            "received": bool(r and r.received) if r else False,
            "requested": bool(r and r.requested) if r else False,
            "none_status": bool(r and r.none_status) if r else False,
            "comment": (r.comment or "") if r else "",
        })

    # 관세법인 드림만 입력 가능, 환경산업기술원·지원기업은 읽기 전용
    can_edit = role in (UserRole.DREAM, UserRole.ADMIN)

    st.markdown("**사전진단체크리스트** (양식: 사전진단체크리스트.xlsx 기준)")

    # CSS: 표 형태의 행·셀 스타일
    st.markdown("""
    <style>
    div[data-testid="column"] .stTextInput input { font-size: 0.9em; }
    .prediag-header-row { font-weight: 600; background-color: #f0f0f0; padding: 8px 6px; border: 1px solid #ccc; border-bottom: none; }
    .prediag-cell-section { background-color: #e8e8e8; padding: 6px 8px; border: 1px solid #ccc; font-size: 0.9em; min-height: 2.2em; }
    .prediag-cell { padding: 4px 8px; border: 1px solid #ccc; font-size: 0.9em; min-height: 2.2em; }
    </style>
    """, unsafe_allow_html=True)

    # 헤더 행
    h_cols = st.columns([1, 1.2, 2, 1.2, 0.5, 0.5, 0.5, 1.5, 2])
    headers = ["구분", "항목", "내용", "관련 서류", "수취", "요청", "없음", "비고/가이드", "Comment"]
    for i, h in enumerate(headers):
        with h_cols[i]:
            st.markdown(f"**{h}**")

    # 데이터 행: 표 안에서 내용·Comment는 입력란, 수취/요청/없음은 체크박스 (편집 권한 시)
    for row in display_rows:
        tid = row["template_id"]
        cols = st.columns([1, 1.2, 2, 1.2, 0.5, 0.5, 0.5, 1.5, 2])
        with cols[0]:
            if row["section_rowspan"] > 0:
                st.markdown(f'<div class="prediag-cell-section">{row["section"]}</div>', unsafe_allow_html=True)
            else:
                st.write("")
        with cols[1]:
            st.caption(row["item"])
        with cols[2]:
            if can_edit:
                st.text_input("내용", value=row["content"], key=f"pre_diag_content_{tid}", label_visibility="collapsed")
            else:
                st.caption(row["content"] or "")
        with cols[3]:
            st.caption(row["related_doc"] or "")
        with cols[4]:
            if can_edit:
                st.checkbox("수취", value=row["received"], key=f"pre_diag_rec_{tid}", label_visibility="collapsed")
            else:
                st.caption("✓" if row["received"] else "")
        with cols[5]:
            if can_edit:
                st.checkbox("요청", value=row["requested"], key=f"pre_diag_req_{tid}", label_visibility="collapsed")
            else:
                st.caption("✓" if row["requested"] else "")
        with cols[6]:
            if can_edit:
                st.checkbox("없음", value=row["none_status"], key=f"pre_diag_non_{tid}", label_visibility="collapsed")
            else:
                st.caption("✓" if row["none_status"] else "")
        with cols[7]:
            st.caption((row["guide_notes"] or "")[:120])
        with cols[8]:
            if can_edit:
                st.text_input("Comment", value=row["comment"], key=f"pre_diag_comment_{tid}", label_visibility="collapsed")
            else:
                st.caption(row["comment"] or "")

    if can_edit:
        if st.button("체크리스트 저장", key="pre_diag_save"):
            for row in display_rows:
                tid = row["template_id"]
                content = st.session_state.get(f"pre_diag_content_{tid}", "")
                rec = st.session_state.get(f"pre_diag_rec_{tid}", False)
                req = st.session_state.get(f"pre_diag_req_{tid}", False)
                non = st.session_state.get(f"pre_diag_non_{tid}", False)
                comment = st.session_state.get(f"pre_diag_comment_{tid}", "")
                resp = responses_by_tid.get(tid)
                if resp is None:
                    resp = PreDiagnosisChecklistResponse(company_id=company.id, template_id=tid)
                    session.add(resp)
                resp.content = str(content).strip() or None
                resp.received = 1 if rec else 0
                resp.requested = 1 if req else 0
                resp.none_status = 1 if non else 0
                resp.comment = str(comment).strip() or None
            session.commit()
            log_history(session, company.id, "사전진단체크리스트 등록/수정", "체크리스트 저장")
            st.success("체크리스트가 저장되었습니다.")
            st.rerun()

    # ---- 사전진단체크리스트 첨부자료 ----
    st.markdown("---")
    st.markdown("#### 사전진단체크리스트 첨부자료")
    attachments = (
        session.query(PreDiagnosisAttachment)
        .filter(PreDiagnosisAttachment.company_id == company.id)
        .order_by(PreDiagnosisAttachment.id)
        .all()
    )
    if can_edit:
        with st.expander("첨부자료 추가"):
            with st.form("pre_diag_attachment_form"):
                title = st.text_input("자료 제목", key="pre_diag_attach_title")
                file = st.file_uploader("파일", type=["pdf", "xlsx", "xls", "docx", "doc"], key="pre_diag_attach_file")
                if st.form_submit_button("첨부 저장"):
                    if title and str(title).strip():
                        path = save_uploaded_file(company.id, "pre_diag_attach", file) if file else None
                        session.add(PreDiagnosisAttachment(company_id=company.id, title=str(title).strip(), file_path=path))
                        session.commit()
                        log_history(session, company.id, "사전진단체크리스트 등록/수정", "첨부자료 추가")
                        st.success("첨부자료가 저장되었습니다.")
                        st.rerun()
                    else:
                        st.warning("제목을 입력해 주세요.")
    if attachments:
        for a in attachments:
            st.write(f"- **{a.title}** — {a.file_path or '(파일 없음)'}")
    else:
        st.info("등록된 첨부자료가 없습니다.")

    st.markdown("---")
    # ---- 수입국 통상환경 분석 보고서 (정형 항목) ----
    st.markdown("#### 수입국 통상환경 분석 보고서 (정형 입력)")
    latest_env = (
        session.query(TradeEnvironmentReport)
        .filter(TradeEnvironmentReport.company_id == company.id)
        .order_by(TradeEnvironmentReport.id.desc())
        .first()
    )
    if role in (UserRole.ADMIN, UserRole.DREAM):
        with st.form("trade_env_form"):
            hs_code_validity = st.text_area(
                "1) 해당 HS code 유효성 검토",
                value=latest_env.hs_code_validity if latest_env and latest_env.hs_code_validity else "",
            )
            hs_detail_criteria = st.text_area(
                "2) 해당 국가 HS code 세부 분류 기준 검토 결과",
                value=latest_env.hs_detail_criteria if latest_env and latest_env.hs_detail_criteria else "",
            )
            import_constraints = st.text_area(
                "3) 수입국 통관 상 제약사항 검토 결과",
                value=latest_env.import_constraints if latest_env and latest_env.import_constraints else "",
            )
            post_import_constraints = st.text_area(
                "4) 수입국 통관 후 제약사항 검토 결과",
                value=latest_env.post_import_constraints if latest_env and latest_env.post_import_constraints else "",
            )
            duty_cert_requirements = st.text_area(
                "5) 기타 관세납부, 인증, 요건 관련 검토 결과",
                value=latest_env.duty_cert_requirements if latest_env and latest_env.duty_cert_requirements else "",
            )
            logistics_considerations = st.text_area(
                "6) 기타 물류 상 주요 고려사항 검토 결과",
                value=latest_env.logistics_considerations if latest_env and latest_env.logistics_considerations else "",
            )
            if st.form_submit_button("통상환경 분석 내용 저장"):
                report = TradeEnvironmentReport(
                    company_id=company.id,
                    hs_code_validity=hs_code_validity.strip() or None,
                    hs_detail_criteria=hs_detail_criteria.strip() or None,
                    import_constraints=import_constraints.strip() or None,
                    post_import_constraints=post_import_constraints.strip() or None,
                    duty_cert_requirements=duty_cert_requirements.strip() or None,
                    logistics_considerations=logistics_considerations.strip() or None,
                )
                session.add(report)
                session.commit()
                process = "수입국 통관절차 검토보고서 등록" if not latest_env else "수입국 통관절차 검토보고서 수정"
                log_history(session, company.id, process, "수입국 통관/관세·인증·물류 검토 내용 업데이트")
                st.success("수입국 통상환경 분석 보고서가 저장되었습니다.")
                st.rerun()

    if latest_env:
        st.markdown("**최근 수입국 통상환경 분석 요약**")
        st.write("1) HS code 유효성 검토:", latest_env.hs_code_validity or "-")
        st.write("2) 국가별 HS code 세부 분류 기준:", latest_env.hs_detail_criteria or "-")
        st.write("3) 수입국 통관 상 제약사항:", latest_env.import_constraints or "-")
        st.write("4) 수입국 통관 후 제약사항:", latest_env.post_import_constraints or "-")
        st.write("5) 관세납부·인증·요건 관련:", latest_env.duty_cert_requirements or "-")
        st.write("6) 물류 상 주요 고려사항:", latest_env.logistics_considerations or "-")
    else:
        st.info("등록된 수입국 통상환경 분석 보고서가 없습니다.")

    st.markdown("---")
    # ---- 기존 파일 업로드 (체크리스트 파일, 통상환경 보고서, 관세실익분석표) ----
    pa = (
        session.query(PreAnalysis)
        .filter(PreAnalysis.company_id == company.id)
        .order_by(PreAnalysis.id.desc())
        .first()
    )
    if role in (UserRole.ADMIN, UserRole.DREAM):
        with st.expander("사전진단 자료 파일 업로드 (체크리스트·통상환경·관세실익)"):
            with st.form("pre_analysis_upload"):
                c1 = st.file_uploader("체크리스트 파일", type=["pdf", "xlsx", "docx"], key="pre_check")
                c2 = st.file_uploader("수입국 통상환경 보고서", type=["pdf", "xlsx", "docx"], key="pre_trade")
                c3 = st.file_uploader("관세실익분석표", type=["pdf", "xlsx", "docx"], key="pre_benefit")
                notes = st.text_area("메모", value=pa.notes if pa and pa.notes else "")
                if st.form_submit_button("파일 저장"):
                    pa_new = PreAnalysis(company_id=company.id, notes=notes.strip() or None)
                    pa_new.checklist_path = save_uploaded_file(company.id, "pre_analysis", c1) or (pa.checklist_path if pa else None)
                    pa_new.trade_report_path = save_uploaded_file(company.id, "pre_analysis", c2) or (pa.trade_report_path if pa else None)
                    pa_new.benefit_analysis_path = save_uploaded_file(company.id, "pre_analysis", c3) or (pa.benefit_analysis_path if pa else None)
                    session.add(pa_new)
                    session.commit()
                    process = "사전진단체크리스트 등록" if not pa else "사전진단체크리스트 수정"
                    log_history(session, company.id, process, "사전진단체크리스트/통상환경/관세실익 자료 업데이트")
                    st.success("저장되었습니다.")
                    st.rerun()
    st.markdown("**첨부 파일 현황**")
    if pa and (pa.checklist_path or pa.trade_report_path or pa.benefit_analysis_path):
        st.write("- 체크리스트:", pa.checklist_path or "-")
        st.write("- 수입국 통상환경 보고서:", pa.trade_report_path or "-")
        st.write("- 관세실익분석표:", pa.benefit_analysis_path or "-")
        if pa.notes:
            st.write("- 메모:", pa.notes)
    else:
        st.info("등록된 첨부 파일이 없습니다.")


def page_hs_code(
    session: Session, company: Optional[Company], role: UserRole
):
    st.subheader("HS code관리 - 품목분류 결과보고서")
    if not company:
        return
    latest = (
        session.query(HSCodeReport)
        .filter(HSCodeReport.company_id == company.id)
        .order_by(HSCodeReport.id.desc())
        .first()
    )
    if role in (UserRole.ADMIN, UserRole.DREAM):
        with st.form("hs_code_upload"):
            f = st.file_uploader("품목분류 결과보고서 업로드", type=["pdf", "xlsx", "xls", "docx"], key="hs_code_file")
            notes = st.text_area("비고", value=latest.notes if latest and latest.notes else "", key="hs_notes")
            if st.form_submit_button("저장"):
                path = save_uploaded_file(company.id, "hs_code", f)
                r = HSCodeReport(company_id=company.id, file_path=path or latest.file_path if latest else None, notes=notes.strip() or None)
                session.add(r)
                session.commit()
                process = "품목분류검토보고서 등록" if not latest else "품목분류검토보고서 수정"
                log_history(session, company.id, process, "HS code 검토보고서 업데이트")
                st.success("저장되었습니다.")
                st.rerun()
    st.markdown("---")
    st.markdown("**진행상황 / 최근 보고서**")
    if latest:
        st.write("파일:", latest.file_path or "-")
        st.write("메모:", latest.notes or "-")
    else:
        st.info("등록된 보고서가 없습니다.")


def page_execution_plan(session: Session, company: Optional[Company], role: UserRole):
    st.subheader("수행계획서")
    if not company:
        return
    latest = (
        session.query(ExecutionPlan)
        .filter(ExecutionPlan.company_id == company.id)
        .order_by(ExecutionPlan.id.desc())
        .first()
    )
    if role in (UserRole.ADMIN, UserRole.DREAM):
        with st.form("exec_plan"):
            f = st.file_uploader("수행계획서 파일", type=["pdf", "xlsx", "xls", "docx"], key="exec_plan_file")
            notes = st.text_area("비고", value=latest.notes if latest and latest.notes else "", key="exec_notes")
            if st.form_submit_button("업로드/저장"):
                path = save_uploaded_file(company.id, "execution_plan", f)
                ep = ExecutionPlan(company_id=company.id, plan_file_path=path or (latest.plan_file_path if latest else None), notes=notes.strip() or None)
                session.add(ep)
                session.commit()
                process = "수행계획서 등록" if not latest else "수행계획서 수정"
                log_history(session, company.id, process, "수행계획서가 업데이트되었습니다.")
                st.success("저장되었습니다.")
                st.rerun()
    st.markdown("---")
    if latest and latest.plan_file_path:
        st.write("최근 파일:", latest.plan_file_path)
    else:
        st.info("등록된 수행계획서가 없습니다.")


def page_doc_checklist(
    session: Session, company: Optional[Company], role: UserRole
):
    st.subheader("서류준비 체크리스트 - 서류 준비 상황 확인")
    if not company:
        return
    if role in (UserRole.ADMIN, UserRole.DREAM):
        st.markdown("#### 서류준비 완료 처리")
        st.caption("모든 서류 준비가 완료되었을 때 '완료'를 눌러 진행단계 보드에 반영합니다.")
        col_a, col_b, col_c = st.columns([1, 1, 2])
        with col_a:
            if st.button("서류준비 완료", key=f"doc_done_{company.id}"):
                company.doc_checklist_completed = 1
                company.doc_checklist_completed_at = dt_util.utcnow()
                session.commit()
                log_history(session, company.id, "서류준비 체크리스트 업데이트", "서류준비 체크리스트 완료 처리")
                st.success("서류준비가 완료 처리되었습니다.")
                st.rerun()
        with col_b:
            if company.doc_checklist_completed:
                st.markdown(
                    '<span style="color:red; font-weight:600;">서류준비 완료 취소</span>',
                    unsafe_allow_html=True,
                )
                if st.button("서류준비 완료 취소", key=f"doc_done_cancel_{company.id}"):
                    company.doc_checklist_completed = 0
                    company.doc_checklist_completed_at = None
                    session.commit()
                    log_history(session, company.id, "서류준비 체크리스트 업데이트", "서류준비 체크리스트 완료 취소")
                    st.success("서류준비 완료가 취소되었습니다.")
                    st.rerun()
        with col_c:
            st.write("현재 상태:", "완료" if company.doc_checklist_completed else "미완료")
        st.markdown("---")
        if company.doc_checklist_completed:
            st.warning("서류준비가 완료 상태입니다. 추가 요청서류 등록이나 자료 수정이 필요하면 '서류준비 완료 취소' 후 진행하는 것을 권장합니다.")

    done = bool(company.doc_checklist_completed)
    # 기본 항목 + 해당 기업 추가 항목
    default_items = session.query(DocumentItem).filter(DocumentItem.company_id.is_(None)).order_by(DocumentItem.sort_order).all()
    custom_items = session.query(DocumentItem).filter(DocumentItem.company_id == company.id).order_by(DocumentItem.sort_order).all()
    all_items = list(default_items) + list(custom_items)

    if role in (UserRole.ADMIN, UserRole.DREAM) and not done:
        with st.expander("관세법인: 추가 요청서류 등록"):
            with st.form("add_doc_item"):
                new_name = st.text_input("서류명")
                if st.form_submit_button("추가"):
                    if new_name.strip():
                        max_order = max((i.sort_order for i in custom_items), default=len(default_items))
                        session.add(DocumentItem(company_id=company.id, name=new_name.strip(), sort_order=max_order + 1))
                        session.commit()
                        st.success("추가되었습니다.")
                        st.rerun()

    for item in all_items:
        status = (
            session.query(DocChecklistStatus)
            .filter(
                DocChecklistStatus.company_id == company.id,
                DocChecklistStatus.document_item_id == item.id,
            )
            .first()
        )
        col1, col2, col3 = st.columns([2, 1, 2])
        with col1:
            st.write(item.name)
        with col2:
            checked = st.checkbox(
                "완료",
                value=bool(status and status.is_checked),
                key=f"doc_check_{company.id}_{item.id}",
            )
        with col3:
            if role in (UserRole.ADMIN, UserRole.DREAM) and not done:
                up = st.file_uploader("파일", key=f"doc_file_{company.id}_{item.id}", type=["pdf", "xlsx", "docx"])
                if st.button("저장", key=f"doc_save_{company.id}_{item.id}"):
                    path = save_uploaded_file(company.id, "doc_checklist", up) if up is not None else (status.file_path if status else None)
                    # 파일이 새로 업로드되었거나 기존 체크 상태가 유지/변경되었을 때 완료 여부 결정
                    new_checked = bool(checked or up)
                    if status:
                        status.is_checked = 1 if new_checked else 0
                        status.file_path = path
                    else:
                        session.add(
                            DocChecklistStatus(
                                company_id=company.id,
                                document_item_id=item.id,
                                is_checked=1 if new_checked else 0,
                                file_path=path,
                            )
                        )
                    session.commit()
                    if up is not None and path:
                        log_history(session, company.id, "서류준비 체크리스트 업데이트", f"{item.name} 업로드 완료")
                    st.success("저장됨")
                    st.rerun()
            else:
                # 완료 상태이거나 권한이 없는 경우: 업로드/저장 불가, 파일 경로만 표시
                if status and status.file_path:
                    st.caption(f"파일: {status.file_path}")
        st.markdown("---")


def page_delivery_timeline(
    session: Session, company: Optional[Company], role: UserRole
):
    st.subheader("배송일정 타임라인")
    if not company:
        return
    steps = list(ExportStep)
    # 각 단계별 이벤트 (회사당 단계당 1행)
    events = {
        (e.stage): e
        for e in session.query(DeliveryTimelineEvent)
        .filter(DeliveryTimelineEvent.company_id == company.id)
    }
    from datetime import datetime as dt
    for step in steps:
        ev = events.get(step)
        with st.expander(f"{step.label} — 계획: {ev.planned_date if ev and ev.planned_date else '-'} / 완료: {ev.actual_completed_at if ev and ev.actual_completed_at else '-'}"):
            if role in (UserRole.ADMIN, UserRole.DREAM):
                with st.form(f"timeline_{step.name}"):
                    planned = st.date_input("계획일", value=ev.planned_date.date() if ev and ev.planned_date else None, key=f"plan_{step.name}")
                    actual_d = st.date_input("완료일", value=ev.actual_completed_at.date() if ev and ev.actual_completed_at else None, key=f"act_d_{step.name}")
                    actual_t = st.time_input("완료 시각", value=ev.actual_completed_at.time() if ev and ev.actual_completed_at else dt.now().time(), key=f"act_t_{step.name}")
                    notes = st.text_area("메모", value=ev.notes if ev and ev.notes else "", key=f"notes_{step.name}")
                    if st.form_submit_button("저장"):
                        planned_dt = dt.combine(planned, dt.min.time()) if planned else None
                        actual_dt = dt.combine(actual_d, actual_t) if actual_d else None
                        if ev:
                            ev.planned_date = planned_dt
                            ev.actual_completed_at = actual_dt
                            ev.notes = notes.strip() or None
                        else:
                            session.add(DeliveryTimelineEvent(company_id=company.id, stage=step, planned_date=planned_dt, actual_completed_at=actual_dt, notes=notes.strip() or None))
                        session.commit()
                        log_history(session, company.id, "배송일정 타임라인 업데이트", f"{step.label} 단계 일정/완료 상태가 업데이트되었습니다.")
                        st.success("저장됨")
                        st.rerun()
            else:
                if ev:
                    st.write("계획일:", ev.planned_date)
                    st.write("완료일시:", ev.actual_completed_at)
                    if ev.notes:
                        st.write("메모:", ev.notes)


def page_stage_board(session: Session, companies: List[Company]):
    st.markdown('<p class="section-title">📋 전체 진행단계별 보드</p>', unsafe_allow_html=True)
    if not companies:
        st.info("조회 가능한 기업이 없습니다.")
        return
    # 요구사항 기반 헤더로 변경
    cols = [
        "업체명",
        "담당 관세사",
        "등록일자",
        "배정일자",
        "사전진단",
        "HS code확정",
        "수행계획서",
        "서류준비",
        "배송준비",
        "현지통관",
        "배송완료",
    ]
    data = []
    for c in companies:
        if not c.assigned_consultant:
            continue
        ps = company_progress_status(session, c)
        # 배송준비/현지통관은 기존 타임라인 이벤트를 참고해 보조적으로 표시(없으면 '-')
        evs = session.query(DeliveryTimelineEvent).filter(DeliveryTimelineEvent.company_id == c.id).all()
        by_stage = {e.stage: e for e in evs}
        shipping_prep = "완료" if (by_stage.get(ExportStep.PRE_SHIPMENT) and by_stage.get(ExportStep.PRE_SHIPMENT).actual_completed_at) else "-"
        local_clearance = "완료" if (by_stage.get(ExportStep.IMPORT_CLEARANCE) and by_stage.get(ExportStep.IMPORT_CLEARANCE).actual_completed_at) else "-"
        data.append({
            "업체명": c.name,
            "담당 관세사": c.assigned_consultant or "",
            "등록일자": c.created_at.date().isoformat() if c.created_at else "",
            "배정일자": c.assigned_at.date().isoformat() if c.assigned_at else "",
            "사전진단": "완료" if ps["prediag_done"] else "-",
            "HS code확정": "완료" if ps["hs_done"] else "-",
            "수행계획서": "완료" if ps["exec_done"] else "-",
            "서류준비": "완료" if ps["docs_done"] else "-",
            "배송준비": shipping_prep,
            "현지통관": local_clearance,
            "배송완료": "완료" if ps["result_done"] else "-",
        })
    df = pd.DataFrame(data, columns=cols)
    st.dataframe(df, use_container_width=True, hide_index=True)


def page_signup(session: Session) -> None:
    """회원가입: 구분(KEITI/지원기업/관세법인 드림). 지원기업은 한국환경산업기술원 등록 업체 목록에서 선택."""
    st.subheader("회원가입")
    st.caption("가입 후 관리자 승인을 받아야 로그인할 수 있습니다.")
    with st.form("signup"):
        role_choice = st.selectbox(
            "구분",
            [UserRole.KEITI.label, UserRole.COMPANY.label, UserRole.DREAM.label],
            key="signup_role",
        )
        role = next(r for r in (UserRole.KEITI, UserRole.COMPANY, UserRole.DREAM) if r.label == role_choice)
        # 지원기업일 때만: 한국환경산업기술원이 등록한 지원기업 목록 드롭다운
        company_id_selected = None
        company_name = ""
        if role == UserRole.COMPANY:
            companies = session.query(Company).order_by(Company.name).all()
            if not companies:
                st.warning("등록된 지원기업이 없습니다. 한국환경산업기술원에서 먼저 지원기업을 등록해 주세요.")
            else:
                company_names = [c.name for c in companies]
                company_name = st.selectbox("업체명 (등록된 지원기업 목록)", company_names, key="signup_company")
                company_id_selected = next(c.id for c in companies if c.name == company_name)
        else:
            company_name = st.text_input("업체명", key="signup_company_text")
        display_name = st.text_input("이름")
        phone = st.text_input("전화번호")
        email = st.text_input("메일주소")
        login_id = st.text_input("아이디")
        password = st.text_input("비밀번호", type="password")
        if st.form_submit_button("가입 신청"):
            if not login_id.strip() or not password:
                st.error("아이디와 비밀번호를 입력하세요.")
            elif role == UserRole.COMPANY and not company_id_selected and not company_name:
                st.error("업체를 선택하세요.")
            else:
                existing = session.query(User).filter(User.login_id == login_id.strip()).first()
                if existing:
                    st.error("이미 사용 중인 아이디입니다.")
                else:
                    u = User(
                        login_id=login_id.strip(),
                        password_hash=hash_password(password),
                        role=role,
                        is_approved=0,
                        company_id=company_id_selected,
                        company_name=(company_name.strip() if isinstance(company_name, str) else company_name) or None,
                        display_name=display_name.strip() or None,
                        phone=phone.strip() or None,
                        email=email.strip() or None,
                    )
                    session.add(u)
                    session.commit()
                    st.success("회원가입 신청이 완료되었습니다. 관리자 승인 후 로그인할 수 있습니다.")
                    if st.session_state.get("show_signup"):
                        st.session_state.pop("show_signup", None)
                    st.rerun()
    if st.button("로그인 화면으로"):
        st.session_state.pop("show_signup", None)
        st.rerun()


def page_company_register(session: Session) -> None:
    """한국환경산업기술원: 지원기업 등록 — 기업명 및 각 정보별 입력."""
    st.subheader("지원기업 등록")
    st.caption("한국환경산업기술원에서 지원기업으로 등록한 업체 목록이 담당 관세사 배정·회원가입 시 드롭다운에 사용됩니다.")
    with st.form("company_register"):
        name = st.text_input("기업명 *")
        application_info = st.text_area("신청 정보")
        contact_person = st.text_input("담당자")
        contact_phone = st.text_input("연락처")
        contact_email = st.text_input("이메일")
        address = st.text_area("주소")
        if st.form_submit_button("등록"):
            if not name.strip():
                st.error("기업명을 입력하세요.")
            else:
                existing = session.query(Company).filter(Company.name == name.strip()).first()
                if existing:
                    st.error("이미 등록된 기업명입니다.")
                else:
                    c = Company(
                        name=name.strip(),
                        application_info=application_info.strip() or None,
                        contact_person=contact_person.strip() or None,
                        contact_phone=contact_phone.strip() or None,
                        contact_email=contact_email.strip() or None,
                        address=address.strip() or None,
                    )
                    session.add(c)
                    session.commit()
                    log_history(session, c.id, "지원기업등록", f"지원기업 등록 ({c.name})")
                    st.success(f"지원기업이 등록되었습니다: {c.name}")
                    st.rerun()
    st.markdown("---")
    st.markdown("**등록된 지원기업 목록**")
    companies = session.query(Company).order_by(Company.name).all()
    if not companies:
        st.info("등록된 지원기업이 없습니다.")
    else:
        for c in companies:
            st.markdown(f"- **{c.name}** | 담당자: {c.contact_person or '-'} | 연락처: {c.contact_phone or '-'} | 담당 관세사: {c.assigned_consultant or '미배정'}")


def page_consultant_register(session: Session) -> None:
    """관세법인 드림: 관세사 등록 — 담당 관세사 배정 시 드롭다운에 사용."""
    st.subheader("관세사 등록")
    st.caption("등록한 관세사는 담당 관세사 배정 메뉴에서 드롭다운으로 선택할 수 있습니다.")
    with st.form("consultant_register"):
        consultant_name = st.text_input("관세사 이름 *")
        if st.form_submit_button("등록"):
            if not consultant_name.strip():
                st.error("관세사 이름을 입력하세요.")
            else:
                existing = session.query(Consultant).filter(Consultant.name == consultant_name.strip()).first()
                if existing:
                    st.error("이미 등록된 관세사입니다.")
                else:
                    session.add(Consultant(name=consultant_name.strip()))
                    session.commit()
                    st.success(f"관세사가 등록되었습니다: {consultant_name.strip()}")
                    st.rerun()
    st.markdown("---")
    st.markdown("**등록된 관세사 목록**")
    consultants = session.query(Consultant).order_by(Consultant.name).all()
    if not consultants:
        st.info("등록된 관세사가 없습니다.")
    else:
        for c in consultants:
            st.markdown(f"- {c.name}")


def page_assign_consultant(session: Session, companies: List[Company]) -> None:
    st.subheader("담당 관세사 배정")
    if not companies:
        st.info("등록된 지원기업이 없습니다.")
        return
    with st.form("assign_consultant_main"):
        ac_company = st.selectbox("기업 (지원기업 등록 목록)", [c.name for c in companies], key="ac_company_main")
        consultants = session.query(Consultant).order_by(Consultant.name).all()
        if consultants:
            consultant = st.selectbox("담당 관세사", [c.name for c in consultants], key="ac_consultant_main")
        else:
            st.caption("관세사 등록 메뉴에서 관세사를 먼저 등록하세요.")
            consultant = st.text_input("담당 관세사 이름 (직접 입력)", key="ac_consultant_text_main") or ""
        if st.form_submit_button("배정 저장"):
            comp = next(c for c in companies if c.name == ac_company)
            comp.assigned_consultant = (consultant.strip() if isinstance(consultant, str) else consultant) or None
            if not comp.assigned_at:
                comp.assigned_at = dt_util.utcnow()
            session.commit()
            log_history(
                session,
                comp.id,
                "담당 관세사 배정",
                f"담당 관세사 '{comp.assigned_consultant}' 배정",
            )
            st.success("담당 관세사가 배정되었습니다.")
            st.rerun()


def page_result_report(session: Session, company: Optional[Company], role: UserRole) -> None:
    st.subheader("결과보고서")
    if not company:
        return
    latest = (
        session.query(ResultReport)
        .filter(ResultReport.company_id == company.id)
        .order_by(ResultReport.id.desc())
        .first()
    )
    if role in (UserRole.ADMIN, UserRole.DREAM):
        with st.form("result_report_upload"):
            f = st.file_uploader("결과보고서 업로드", type=["pdf", "xlsx", "xls", "docx"], key="result_report_file")
            notes = st.text_area("비고", value=latest.notes if latest and latest.notes else "")
            if st.form_submit_button("저장"):
                path = save_uploaded_file(company.id, "result_report", f) if f else (latest.file_path if latest else None)
                rr = ResultReport(company_id=company.id, file_path=path, notes=notes.strip() or None)
                session.add(rr)
                session.commit()
                log_history(session, company.id, "결과보고서 등록", "결과보고서가 등록되었습니다.")
                st.success("저장되었습니다.")
                st.rerun()
    st.markdown("---")
    if latest:
        st.write("파일:", latest.file_path or "-")
        st.write("비고:", latest.notes or "-")
    else:
        st.info("등록된 결과보고서가 없습니다.")


def page_history(session: Session, companies: List[Company], company: Optional[Company], role: UserRole) -> None:
    st.subheader("히스토리 관리")
    if not companies:
        st.info("조회 가능한 기업이 없습니다.")
        return
    if not company:
        st.info("상단에서 기업을 선택하세요.")
        return

    st.markdown(f"**업체명:** {company.name}  \n**담당 관세사:** {company.assigned_consultant or '-'}")

    st.markdown("---")
    st.markdown("### 히스토리 타임라인")
    # 오래된 이벤트부터 조회 후, 화면에도 오래된 것 → 최신 순으로 위에서 아래로 표시
    entries = (
        session.query(HistoryEntry)
        .filter(HistoryEntry.company_id == company.id)
        .order_by(HistoryEntry.created_at.asc())
        .all()
    )
    if not entries:
        st.info("등록된 히스토리가 없습니다.")
        return

    # 오래된 것 → 최신 순으로, 배송 추적/채팅 스타일 타임라인 표시
    st.markdown('<div class="history-timeline">', unsafe_allow_html=True)
    for e in entries:
        ts = e.created_at.strftime("%Y-%m-%d %H:%M")
        actor = e.actor or "작업자 미기록"
        html = f"""
        <div class="history-item">
            <div class="history-dot"></div>
            <div class="history-content">
                <div class="history-title">{e.process}</div>
                <div class="history-meta">{ts} · {actor}</div>
                <div class="history-desc">{e.status}</div>
            </div>
        </div>
        """
        st.markdown(html, unsafe_allow_html=True)
    st.markdown("</div>", unsafe_allow_html=True)


def page_company_detail(session: Session, company: Optional[Company], role: UserRole) -> None:
    st.subheader("기업 상세")
    if not company:
        st.info("기업을 선택하세요.")
        return
    ps = company_progress_status(session, company)
    st.markdown(f"**업체명:** {company.name}  \n**담당 관세사:** {company.assigned_consultant or '-'}  \n**현재 단계:** {ps['current']}")
    tabs = st.tabs(["사전진단", "HS code", "수행계획서", "서류준비", "배송일정", "결과보고서"])
    with tabs[0]:
        page_pre_diagnosis(session, company, role)
    with tabs[1]:
        page_hs_code(session, company, role)
    with tabs[2]:
        page_execution_plan(session, company, role)
    with tabs[3]:
        page_doc_checklist(session, company, role)
    with tabs[4]:
        page_delivery_timeline(session, company, role)
    with tabs[5]:
        page_result_report(session, company, role)


def page_my_profile(session: Session, user: User) -> None:
    """로그인한 사용자의 개인정보를 수정하는 화면. 비밀번호는 즉시 반영, 그 외 항목은 관리자 승인 후 반영."""
    st.subheader("개인정보 설정")
    st.caption("로그인한 사용자 본인의 정보를 확인하고 수정합니다.")
    st.markdown("**현재 정보**")
    st.write("아이디:", user.login_id, "(변경 불가)")
    st.write("업체명:", user.company_name or "-")
    st.write("이름:", user.display_name or "-")
    st.write("전화번호:", user.phone or "-")
    st.write("메일주소:", user.email or "-")
    # 승인 대기 중인 수정 요청
    pending = (
        session.query(PendingProfileUpdate)
        .filter(PendingProfileUpdate.user_id == user.id, PendingProfileUpdate.status == "pending")
        .first()
    )
    if pending:
        st.info("회원 정보 수정 요청이 승인 대기 중입니다. 관리자 승인 후 반영됩니다.")
    st.markdown("---")
    with st.form("profile_edit"):
        new_company_name = st.text_input("업체명", value=user.company_name or "")
        new_display_name = st.text_input("이름", value=user.display_name or "")
        new_phone = st.text_input("전화번호", value=user.phone or "")
        new_email = st.text_input("메일주소", value=user.email or "")
        new_password = st.text_input("비밀번호 변경 (변경 시에만 입력)", type="password", placeholder="비워두면 유지")
        if st.form_submit_button("저장"):
            if new_password:
                user.password_hash = hash_password(new_password)
                session.commit()
                st.success("비밀번호가 변경되었습니다.")
            changed = (
                (new_company_name.strip() or "") != (user.company_name or "")
                or (new_display_name.strip() or "") != (user.display_name or "")
                or (new_phone.strip() or "") != (user.phone or "")
                or (new_email.strip() or "") != (user.email or "")
            )
            if changed:
                if pending:
                    st.warning("이미 수정 요청이 대기 중입니다. 승인/거절 후 다시 요청하세요.")
                else:
                    session.add(
                        PendingProfileUpdate(
                            user_id=user.id,
                            requested_company_name=new_company_name.strip() or None,
                            requested_display_name=new_display_name.strip() or None,
                            requested_phone=new_phone.strip() or None,
                            requested_email=new_email.strip() or None,
                            status="pending",
                        )
                    )
                    session.commit()
                    st.success("회원 정보 수정을 요청했습니다. 관리자 승인 후 반영됩니다.")
            if new_password or changed:
                st.rerun()
    st.caption("비밀번호는 즉시 반영됩니다. 업체명·이름·전화번호·메일주소는 관리자 승인 후 반영됩니다.")


def page_admin(session: Session) -> None:
    """관리자 설정: 회원가입 승인, 개인정보 수정 승인."""
    st.subheader("관리자 설정")
    tab1, tab2, tab3 = st.tabs(["회원가입 승인", "개인정보 수정 승인", "회원 목록"])

    with tab1:
        pending_users = session.query(User).filter(User.is_approved != 1).order_by(User.created_at.desc()).all()
        if not pending_users:
            st.info("승인 대기 중인 회원가입이 없습니다.")
        else:
            for u in pending_users:
                with st.expander(f"{u.login_id} | {u.role.label} | {u.company_name or '-'} | {u.display_name or '-'}"):
                    st.write("업체명:", u.company_name or "-")
                    st.write("이름:", u.display_name or "-")
                    st.write("전화번호:", u.phone or "-")
                    st.write("메일주소:", u.email or "-")
                    st.write("가입일:", u.created_at)
                    companies_for_approve = session.query(Company).order_by(Company.name).all()
                    comp_sel = "(미선택)"
                    if u.role == UserRole.COMPANY and companies_for_approve:
                        comp_sel = st.selectbox(
                            "연결할 기업 (지원기업인 경우)",
                            ["(미선택)"] + [c.name for c in companies_for_approve],
                            key=f"approve_comp_{u.id}",
                        )
                    col1, col2 = st.columns(2)
                    with col1:
                        if st.button("승인", key=f"approve_{u.id}"):
                            u.is_approved = 1
                            if u.role == UserRole.COMPANY and comp_sel != "(미선택)" and companies_for_approve:
                                comp = next(c for c in companies_for_approve if c.name == comp_sel)
                                u.company_id = comp.id
                            session.commit()
                            st.success("승인되었습니다.")
                            st.rerun()
                    with col2:
                        if st.button("거절", key=f"reject_{u.id}"):
                            session.delete(u)
                            session.commit()
                            st.success("거절되었습니다.")
                            st.rerun()

    with tab2:
        profile_requests = (
            session.query(PendingProfileUpdate)
            .filter(PendingProfileUpdate.status == "pending")
            .order_by(PendingProfileUpdate.requested_at.desc())
            .all()
        )
        if not profile_requests:
            st.info("승인 대기 중인 회원 정보 수정 요청이 없습니다.")
        else:
            for req in profile_requests:
                usr = session.query(User).filter(User.id == req.user_id).first()
                with st.expander(f"{usr.login_id} | {usr.display_name or '-'} | 요청일 {req.requested_at}"):
                    st.write("요청 업체명:", req.requested_company_name or "-")
                    st.write("요청 이름:", req.requested_display_name or "-")
                    st.write("요청 전화번호:", req.requested_phone or "-")
                    st.write("요청 메일주소:", req.requested_email or "-")
                    col1, col2 = st.columns(2)
                    with col1:
                        if st.button("승인", key=f"prof_approve_{req.id}"):
                            usr.company_name = req.requested_company_name
                            usr.display_name = req.requested_display_name
                            usr.phone = req.requested_phone
                            usr.email = req.requested_email
                            req.status = "approved"
                            req.decided_at = dt_util.utcnow()
                            req.decided_by = st.session_state.get("user_id")
                            session.commit()
                            st.success("반영되었습니다.")
                            st.rerun()
                    with col2:
                        if st.button("거절", key=f"prof_reject_{req.id}"):
                            req.status = "rejected"
                            req.decided_at = dt_util.utcnow()
                            req.decided_by = st.session_state.get("user_id")
                            session.commit()
                            st.success("거절되었습니다.")
                            st.rerun()

    with tab3:
        st.caption("비밀번호는 보안상 평문으로 저장되지 않습니다. 필요 시 아래에서 비밀번호를 초기화하세요.")
        users = session.query(User).order_by(User.created_at.desc()).all()
        if not users:
            st.info("등록된 회원이 없습니다.")
        else:
            rows = []
            for u in users:
                rows.append(
                    {
                        "기업명": u.company_name or "-",
                        "아이디": u.login_id,
                        "구분": u.role.label,
                        "승인": "Y" if u.is_approved == 1 else "N",
                        "이름": u.display_name or "-",
                        "전화번호": u.phone or "-",
                        "메일주소": u.email or "-",
                        "비밀번호(저장)": "해시 저장(평문 조회 불가)",
                    }
                )
            st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

            st.markdown("---")
            st.markdown("#### 비밀번호 초기화")
            approved_users = [u for u in users if u.is_approved == 1]
            if not approved_users:
                st.info("승인된 회원이 없습니다.")
            else:
                sel_login = st.selectbox("계정 선택", [u.login_id for u in approved_users], key="pw_reset_sel")
                new_pw = st.text_input("새 비밀번호", type="password", key="pw_reset_new")
                if st.button("비밀번호 초기화", key="pw_reset_btn"):
                    if not new_pw:
                        st.error("새 비밀번호를 입력하세요.")
                    else:
                        u = next(x for x in approved_users if x.login_id == sel_login)
                        u.password_hash = hash_password(new_pw)
                        session.commit()
                        st.success(f"{sel_login} 비밀번호가 변경되었습니다.")
                        st.rerun()


def inject_custom_css():
    """모니터링 시스템 전용 UI 스타일"""
    st.markdown(
        """
        <style>
        /* 헤더(컬러박스) - 좌측 메뉴바 우측 끝에서 시작, 메뉴바와 함께 이동 */
        .monitoring-header {
            position: fixed;
            top: 3.5rem;
            left: var(--sidebar-width, 21rem);
            right: 0;
            z-index: 999;
            color: #ffffff !important;
            min-height: 72px;
            height: auto;
            padding: 0.75rem 2rem 0.85rem;
            display: flex;
            flex-direction: column;
            justify-content: center;
            align-items: center;
            text-align: center;
            box-shadow: 0 4px 16px rgba(0, 0, 0, 0.2);
            box-sizing: border-box;
            overflow: visible;
            transition: left 0.2s ease;
        }
        /* 사이드바 접힌 경우 컬러박스 시작점(좁은 스트립만 남을 때) */
        [data-testid="stSidebar"][aria-expanded="false"] ~ section.main .monitoring-header {
            left: 2.5rem;
        }
        /* 역할별 헤더 배경색 */
        .monitoring-header.header-role-keiti {
            background: linear-gradient(135deg, #0d5c2e 0%, #1a7d3e 50%, #22a352 100%);
        }
        .monitoring-header.header-role-dream {
            background: linear-gradient(135deg, #374151 0%, #4b5563 50%, #6b7280 100%);
        }
        .monitoring-header.header-role-company {
            background: linear-gradient(135deg, #1a3352 0%, #1e3a5f 25%, #2d5a87 50%, #3d7ab5 100%);
        }
        .monitoring-header.header-role-admin {
            background: linear-gradient(135deg, #7f1d1d 0%, #b91c1c 50%, #dc2626 100%);
        }
        .monitoring-header.header-role-none {
            background: linear-gradient(135deg, #4b5563 0%, #6b7280 100%);
        }
        .monitoring-header-inner {
            position: relative;
            z-index: 1000;
            pointer-events: auto;
        }
        .monitoring-header h1,
        .monitoring-header p {
            color: #ffffff !important;
            margin: 0;
            opacity: 1 !important;
            visibility: visible !important;
            text-shadow: 0 1px 3px rgba(0,0,0,0.35);
            text-align: center;
            position: relative;
            z-index: 1000;
        }
        .monitoring-header h1 {
            font-size: 1.4rem;
            font-weight: 700;
            letter-spacing: -0.02em;
            line-height: 1.35;
        }
        .monitoring-header p {
            margin: 0.2rem 0 0 0;
            font-size: 0.88rem;
            opacity: 0.98 !important;
            font-weight: 500;
        }
        /* 컬러박스(헤더) 아래에서 본문 시작 - 여러 선택자로 적용 */
        .main-content-spacer {
            display: block;
            height: 14rem;
            min-height: 14rem;
            width: 100%;
            flex-shrink: 0;
        }
        section.main > div,
        div[data-testid="stVerticalBlock"] > div:first-child,
        .block-container {
            padding-top: 0 !important;
        }
        /* 히스토리 타임라인 (채팅/배송 추적 스타일) */
        .history-timeline {
            margin-top: 0.5rem;
            padding-left: 0.25rem;
        }
        .history-item {
            position: relative;
            display: flex;
            align-items: flex-start;
            padding: 0.4rem 0 0.4rem 0.35rem;
        }
        .history-item::before {
            content: "";
            position: absolute;
            left: 0.82rem;  /* 동그라미 중심을 지나는 위치 */
            top: 0;
            bottom: 0;
            border-left: 2px solid #e5e7eb;
        }
        .history-dot {
            width: 10px;
            height: 10px;
            border-radius: 999px;
            background-color: #2563eb;
            margin-right: 0.75rem;
            margin-left: 0.15rem;
            flex-shrink: 0;
            box-shadow: 0 0 0 3px rgba(37, 99, 235, 0.18);
        }
        .history-content {
            flex: 1;
        }
        .history-title {
            font-weight: 600;
            font-size: 0.95rem;
            color: #111827;
        }
        .history-meta {
            font-size: 0.75rem;
            color: #6b7280;
        }
        .history-desc {
            margin-top: 0.1rem;
            font-size: 0.86rem;
            color: #374151;
        }
        .history-item:last-child::before {
            bottom: 0.5rem;
        }

        /* 대시보드 진행상태 표 */
        .progress-table {
            width: 100%;
            border-collapse: collapse;
            margin-top: 0.5rem;
            font-size: 0.85rem;
        }
        .progress-table th,
        .progress-table td {
            padding: 0.45rem 0.5rem;
            border-bottom: 1px solid #e5e7eb;
            text-align: center;
            white-space: nowrap;
        }
        .progress-table th:first-child,
        .progress-table td:first-child {
            text-align: left;
        }
        .status-pill {
            display: inline-block;
            padding: 0.1rem 0.55rem;
            border-radius: 999px;
            font-size: 0.78rem;
            font-weight: 500;
        }
        .status-pending {
            background-color: #fee2e2;
            color: #b91c1c;
        }
        .status-done {
            background-color: #dbeafe;
            color: #1d4ed8;
        }
        .status-progress {
            background-color: #fef9c3;
            color: #a16207;
        }
        .status-none {
            background-color: #f3f4f6;
            color: #6b7280;
        }

        /* KPI 카드 */
        .kpi-card {
            background: linear-gradient(145deg, #f8fafc 0%, #f1f5f9 100%);
            border: 1px solid #e2e8f0;
            border-radius: 10px;
            padding: 1rem 1.25rem;
            text-align: center;
            box-shadow: 0 2px 8px rgba(0,0,0,0.06);
        }
        .kpi-card .value {
            font-size: 1.8rem;
            font-weight: 700;
            color: #1e3a5f;
        }
        .kpi-card .label {
            font-size: 0.85rem;
            color: #64748b;
            margin-top: 0.25rem;
        }
        /* 테이블/데이터 영역 */
        div[data-testid="stDataFrame"] {
            border-radius: 8px;
            overflow: hidden;
            box-shadow: 0 2px 8px rgba(0,0,0,0.06);
        }
        /* 사이드바 */
        [data-testid="stSidebar"] {
            background: linear-gradient(180deg, #f8fafc 0%, #f1f5f9 100%);
        }
        /* 네비게이션 메뉴 버튼: 픽토그램 + 텍스트, 리스트 형태 */
        [data-testid="stSidebar"] .stButton > button {
            width: 100%;
            justify-content: flex-start;
            text-align: left;
            padding: 0.6rem 1rem;
            font-weight: 500;
            border: none;
            border-radius: 0;
            border-bottom: 1px solid #e2e8f0;
            background: transparent;
            color: #334155;
            box-shadow: none;
        }
        [data-testid="stSidebar"] .nav-menu-wrap .stButton > button:hover {
            background: #e2e8f0;
            color: #1e3a5f;
        }
        [data-testid="stSidebar"] .nav-menu-title {
            font-size: 0.75rem;
            font-weight: 600;
            color: #64748b;
            text-transform: uppercase;
            letter-spacing: 0.05em;
            padding: 0.5rem 1rem 0.35rem;
            margin-top: 0.5rem;
        }
        /* 섹션 제목 */
        .section-title {
            color: #1e3a5f;
            font-weight: 600;
            margin-bottom: 0.75rem;
            padding-bottom: 0.35rem;
            border-bottom: 2px solid #2d5a87;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def main():
    st.set_page_config(
        page_title="해외실증지원사업 모니터링 시스템",
        layout="wide",
        initial_sidebar_state="expanded",
    )
    inject_custom_css()
    init_db()
    session = get_session()

    if "role" not in st.session_state:
        st.session_state["role"] = None
    if "company_id" not in st.session_state:
        st.session_state["company_id"] = None

    can_proceed = render_login_or_role(session)
    # 역할별 헤더 색상 (로그인 전에는 회색)
    role = st.session_state.get("role")
    role_class = "header-role-none"
    if role == UserRole.KEITI:
        role_class = "header-role-keiti"
    elif role == UserRole.DREAM:
        role_class = "header-role-dream"
    elif role == UserRole.COMPANY:
        role_class = "header-role-company"
    elif role == UserRole.ADMIN:
        role_class = "header-role-admin"

    st.markdown(
        f"""
        <div class="monitoring-header {role_class}">
            <div class="monitoring-header-inner">
                <h1>해외실증지원사업 통관 및 관세 밀착 컨설팅 모니터링 시스템</h1>
                <p>지원기업, 수출단계, 서류준비상황, 배송관련 사항 등 통합 모니터링</p>
            </div>
        </div>
        <div class="main-content-spacer" aria-hidden="true"></div>
        """,
        unsafe_allow_html=True,
    )

    if not can_proceed:
        if st.session_state.get("show_signup"):
            page_signup(session)
        else:
            st.info("로그인 후 이용해 주세요.")
            if st.button("회원가입"):
                st.session_state["show_signup"] = True
                st.rerun()
        session.close()
        return

    role = st.session_state["role"]
    company_id = st.session_state.get("company_id")
    user_id = st.session_state.get("user_id")
    current_user = session.query(User).filter(User.id == user_id).first() if user_id else None
    # 지원기업은 본인 회사만 보여야 함: company_id가 없으면 업체명으로 자동 매핑
    if role == UserRole.COMPANY and current_user:
        if not company_id and current_user.company_name:
            matched = session.query(Company).filter(Company.name == current_user.company_name).first()
            if matched:
                st.session_state["company_id"] = matched.id
                company_id = matched.id
    companies = get_visible_companies(session, role, company_id)

    # 브라우저 뒤로가기를 위해 URL(page/company)에서 현재 화면 복원
    if "page" in st.query_params:
        page_from_qs = st.query_params.get("page")
        if isinstance(page_from_qs, list):
            page_from_qs = page_from_qs[0] if page_from_qs else None
        if page_from_qs:
            st.session_state["menu"] = page_from_qs
    if "company" in st.query_params:
        comp_from_qs = st.query_params.get("company")
        if isinstance(comp_from_qs, list):
            comp_from_qs = comp_from_qs[0] if comp_from_qs else None
        try:
            if comp_from_qs is not None:
                st.session_state["selected_company_id"] = int(comp_from_qs)
        except Exception:
            pass

    st.sidebar.markdown("---")
    st.sidebar.markdown('<p class="nav-menu-title">메뉴</p>', unsafe_allow_html=True)
    # 레벨1: 대시보드 / 진행관리 / 기타관리
    if "menu" not in st.session_state:
        st.session_state["menu"] = MENU_DASHBOARD
    if "l1_progress_open" not in st.session_state:
        st.session_state["l1_progress_open"] = False
    if "l1_misc_open" not in st.session_state:
        st.session_state["l1_misc_open"] = False

    menu = st.session_state["menu"]

    # 진행/기타 하위 메뉴 목록 정의 (복원 시 사용)
    progress_pages = [
        MENU_PRE_DIAG,
        MENU_HS_CODE,
        MENU_EXEC_PLAN,
        MENU_DOC_CHECKLIST,
        MENU_DELIVERY_TIMELINE,
        MENU_STAGE_BOARD,
        MENU_RESULT_REPORT,
        MENU_HISTORY,
    ]
    misc_items_all = [MENU_MY_PROFILE, MENU_CONSULTANT_REGISTER, MENU_ASSIGN_CONSULTANT, MENU_COMPANY_DETAIL]
    if role in (UserRole.ADMIN, UserRole.KEITI):
        misc_items_all.append(MENU_COMPANY_REGISTER)

    # 현재 메뉴가 어느 그룹에 속하는지에 따라 그룹은 항상 펼쳐진 상태로 유지
    if menu in progress_pages:
        st.session_state["l1_progress_open"] = True
    if menu in misc_items_all:
        st.session_state["l1_misc_open"] = True

    # 1) 대시보드 (단독)
    dash_label = f"{MENU_ICONS.get(MENU_DASHBOARD, '📊')}  {MENU_LEVEL1_DASHBOARD}"
    if menu == MENU_DASHBOARD:
        dash_label = f"▸  {dash_label}"
    if st.sidebar.button(dash_label, key="nav_l1_dashboard"):
        set_nav(MENU_DASHBOARD)
        st.rerun()

    # 2) 진행관리 (레벨1 토글 + 레벨2: 사전진단~결과보고서)
    prog_open = st.session_state["l1_progress_open"]
    prog_prefix = "▾" if prog_open else "▸"
    prog_label = f"{prog_prefix}  {MENU_LEVEL1_PROGRESS}"
    if st.sidebar.button(prog_label, key="nav_l1_progress"):
        st.session_state["l1_progress_open"] = not prog_open
        st.rerun()

    if st.session_state["l1_progress_open"]:
        for p in progress_pages:
            icon = MENU_ICONS.get(p, "•")
            label = f"   {icon}  {p}"
            if menu == p:
                label = f"▸ {label}"
            if st.sidebar.button(label, key=f"nav_progress_{p}"):
                set_nav(p)
                st.rerun()

    # 3) 기타관리 (레벨1 토글 + 레벨2: 개인정보 설정, 관세사 등록, 담당 관세사 배정, 기업 상세, 지원기업 등록(KEITI/관리자))
    misc_open = st.session_state["l1_misc_open"]
    misc_prefix = "▾" if misc_open else "▸"
    misc_label = f"{misc_prefix}  {MENU_LEVEL1_MISC}"
    if st.sidebar.button(misc_label, key="nav_l1_misc"):
        st.session_state["l1_misc_open"] = not misc_open
        st.rerun()

    if st.session_state["l1_misc_open"]:
        for p in misc_items_all:
            # 권한 제한: 관세사 등록/담당 관세사 배정은 관세법인 드림/관리자만 의미 있음
            if p in (MENU_CONSULTANT_REGISTER, MENU_ASSIGN_CONSULTANT) and role not in (UserRole.DREAM, UserRole.ADMIN):
                continue
            icon = MENU_ICONS.get(p, "•")
            label = f"   {icon}  {p}"
            if menu == p:
                label = f"▸ {label}"
            if st.sidebar.button(label, key=f"nav_misc_{p}"):
                set_nav(p)
                st.rerun()

    menu = st.session_state["menu"]

    if menu == MENU_MY_PROFILE:
        if current_user:
            page_my_profile(session, current_user)
        else:
            st.warning("사용자 정보를 찾을 수 없습니다.")
        session.close()
        return
    if menu == MENU_COMPANY_REGISTER:
        page_company_register(session)
        session.close()
        return
    if menu == MENU_CONSULTANT_REGISTER:
        page_consultant_register(session)
        session.close()
        return
    if menu == MENU_ASSIGN_CONSULTANT:
        page_assign_consultant(session, companies)
        session.close()
        return
    if menu == MENU_RESULT_REPORT:
        # 기업 선택은 하단 공통 로직에서
        pass
    if menu == MENU_COMPANY_DETAIL:
        selected_id = st.session_state.get("selected_company_id")
        selected_company = session.query(Company).filter(Company.id == selected_id).first() if selected_id else None
        page_company_detail(session, selected_company, role)
        session.close()
        return
    if menu == MENU_ADMIN:
        page_admin(session)
        session.close()
        return

    # 기업 선택 (지원기업이면 1개 고정)
    company = get_selected_company(session, companies, company_id, "company_sel")

    if menu == MENU_DASHBOARD:
        page_dashboard(session, companies)
    elif menu == MENU_PRE_DIAG:
        page_pre_diagnosis(session, company, role)
    elif menu == MENU_HS_CODE:
        page_hs_code(session, company, role)
    elif menu == MENU_EXEC_PLAN:
        page_execution_plan(session, company, role)
    elif menu == MENU_DOC_CHECKLIST:
        page_doc_checklist(session, company, role)
    elif menu == MENU_DELIVERY_TIMELINE:
        page_delivery_timeline(session, company, role)
    elif menu == MENU_STAGE_BOARD:
        page_stage_board(session, companies)
    elif menu == MENU_RESULT_REPORT:
        page_result_report(session, company, role)
    elif menu == MENU_HISTORY:
        page_history(session, companies, company, role)

    # 지원기업 등록은 메뉴 "지원기업 등록"에서 수행 (KEITI/관리자)
    # 회원가입 시 지원기업은 등록된 기업 목록에서 선택, 관리자 승인으로 로그인 가능

    session.close()


if __name__ == "__main__":
    main()
